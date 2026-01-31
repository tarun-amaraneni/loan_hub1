from decimal import Decimal, InvalidOperation
from django.shortcuts import render
from django.http import JsonResponse
from django.db.models import Sum, Value, DecimalField
from django.db.models.functions import Coalesce
from .models import User, Loan, LoanRepayment, generate_unique_repayment_code
from datetime import datetime
from django.db.models.functions import TruncMonth
from django.db.models import Sum, F, Q  # other imports you might already have
from django.db.models.functions import TruncMonth
from django.db.models import Sum
from django.db.models.functions import TruncMonth  # <<-- Add this
from dateutil.relativedelta import relativedelta
from decimal import Decimal
from django.shortcuts import render, get_object_or_404
import calendar
from django.contrib.auth import authenticate, login, logout


def loans_view(request):
    loan_types = ['MTL LOAN', 'FDL LOAN', 'KVP/NSC LOAN']
    loans = []
    user_name = None
    gen_no = None

    if request.method == 'POST':
        gen_no = request.POST.get('gen_no', '').strip()
        loan_id = request.POST.get('loan_id')

        # -----------------------------------------------------
        # DISPLAY LOANS
        # -----------------------------------------------------
        if gen_no:
            user = User.objects.filter(code=gen_no).first()
            user_name = user.name if user else 'Unknown'

            for loan_type in loan_types:
                loan = Loan.objects.filter(
                    gen_no=gen_no,
                    type_of_loan=loan_type,
                    loan_status='Active'
                ).first()

                if loan:
                    repayments = LoanRepayment.objects.filter(loan=loan).aggregate(
                        total_paid_principal=Coalesce(Sum('paid_to_principal'), Value(0), output_field=DecimalField()),
                        total_paid_interest=Coalesce(Sum('paid_to_interest'), Value(0), output_field=DecimalField())
                    )

                    total_paid_principal = Decimal(repayments['total_paid_principal'])
                    total_paid_interest = Decimal(repayments['total_paid_interest'])

                    remaining_principal = Decimal(loan.amount or 0) - total_paid_principal
                    remaining_interest = Decimal(loan.interest or 0) - total_paid_interest

                    # Auto close loan if cleared
                    if remaining_principal <= 0 and remaining_interest <= 0:
                        loan.loan_status = 'Closed'
                        loan.save(update_fields=['loan_status'])
                        continue

                    loans.append({
                        'loan_type': loan_type,
                        'loan': loan,
                        'balance': remaining_principal,
                        'interest': remaining_interest,
                        'loan_code': loan.code,
                    })

        # -----------------------------------------------------
        # PAYMENT UPDATE SECTION
        # -----------------------------------------------------
        if loan_id:
            try:
                loan = Loan.objects.get(id=loan_id)

                def get_decimal(field):
                    try:
                        return Decimal(request.POST.get(field, '0') or '0')
                    except InvalidOperation:
                        return Decimal('0')

                cash  = get_decimal('cash')
                bank1 = get_decimal('bank1')
                bank2 = get_decimal('bank2')
                adj   = get_decimal('adj')

                total_payment = cash + bank1 + bank2 + adj
                original_payment = total_payment

                repayments = LoanRepayment.objects.filter(loan=loan).aggregate(
                    total_paid_principal=Coalesce(Sum('paid_to_principal'), Value(0), output_field=DecimalField()),
                    total_paid_interest=Coalesce(Sum('paid_to_interest'), Value(0), output_field=DecimalField())
                )

                total_paid_principal = Decimal(repayments['total_paid_principal'])
                total_paid_interest = Decimal(repayments['total_paid_interest'])

                remaining_interest = Decimal(loan.interest or 0) - total_paid_interest
                remaining_principal = Decimal(loan.amount or 0) - total_paid_principal

                # FIRST pay Interest
                paid_to_interest = min(remaining_interest, total_payment)
                total_payment -= paid_to_interest

                # THEN pay Principal
                paid_to_principal = min(remaining_principal, total_payment)
                total_payment -= paid_to_principal

                # Repayment Date Handling
                date_str = request.POST.get('date')
                selected_datetime = None
                if date_str:
                    try:
                        selected_datetime = datetime.strptime(date_str, "%Y-%m-%d")
                    except ValueError:
                        pass

                repayment = LoanRepayment.objects.create(
                    loan=loan,
                    total_payment=original_payment,
                    paid_to_interest=paid_to_interest,
                    paid_to_principal=paid_to_principal,
                    payment_mode='mixed',
                    remarks='',
                    cash=cash,
                    bank1=bank1,
                    bank2=bank2,
                    adj=adj,
                    code=generate_unique_repayment_code(),
                    type_of_loan=loan.type_of_loan
                )

                # Override repayment date if selected
                if selected_datetime:
                    repayment.created_at = selected_datetime
                    repayment.save(update_fields=['created_at'])

                # Auto close loan
                if (remaining_interest - paid_to_interest) <= 0 and \
                   (remaining_principal - paid_to_principal) <= 0:
                    loan.loan_status = 'Closed'
                    loan.save(update_fields=['loan_status'])

                return JsonResponse({'status': 'success'})

            except Loan.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': 'Loan not found'}, status=404)
            except Exception as e:
                return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

    # Fill missing loan types
    existing_types = {item['loan_type'] for item in loans}
    for loan_type in loan_types:
        if loan_type not in existing_types:
            loans.append({
                'loan_type': loan_type,
                'loan': None,
                'balance': Decimal('0'),
                'interest': Decimal('0'),
                'loan_code': ''
            })

    return render(request, 'loans.html', {
        'loans': loans,
        'loan_types': loan_types,
        'user_name': user_name,
        'gen_no': gen_no,
    })

from django.db.models import Sum, Value, DecimalField
from django.db.models.functions import Coalesce
from decimal import Decimal, InvalidOperation
def deposits_view(request):
    deposit_types = ['FIXED DEPOSITS', 'THRIFT FUNDS', 'WELFARE COLLECTIONS']
    loans = []
    user_name = None
    gen_no = None

    if request.method == 'POST':
        gen_no = request.POST.get('gen_no', '').strip()
        loan_id = request.POST.get('loan_id')

        # -----------------------------
        # SEARCH BY GEN NO
        # -----------------------------
        if gen_no:
            user = User.objects.filter(code=gen_no).first()
            user_name = user.name if user else 'Unknown'

            for dep_type in deposit_types:

                loan = Loan.objects.filter(
                    gen_no=gen_no,
                    type_of_loan=dep_type,
                    loan_status='Active'
                ).first()

                if loan:
                    repayments = LoanRepayment.objects.filter(loan=loan).aggregate(
                        total_paid_principal=Coalesce(
                            Sum('paid_to_principal'), Value(0), output_field=DecimalField()
                        ),
                        total_paid_interest=Coalesce(
                            Sum('paid_to_interest'), Value(0), output_field=DecimalField()
                        )
                    )

                    total_paid_principal = Decimal(repayments['total_paid_principal'])
                    total_paid_interest = Decimal(repayments['total_paid_interest'])

                    remaining_principal = Decimal(loan.amount or 0) - total_paid_principal
                    remaining_interest = Decimal(loan.interest or 0) - total_paid_interest

                    # ----------------------------------------
                    # AUTO CLOSE IF DEPOSIT IS FULLY CLEARED
                    # ----------------------------------------
                    if remaining_principal <= 0 and remaining_interest <= 0:
                        loan.loan_status = "Closed"
                        loan.save(update_fields=["loan_status"])
                        continue   # Do NOT show cleared deposits

                    # Add only active deposits
                    loans.append({
                        'loan_type': dep_type,
                        'loan': loan,
                        'balance': remaining_principal,
                        'interest': remaining_interest,
                        'loan_code': loan.code,
                        
                    })

        # -----------------------------
        # PAYMENT UPDATE
        # -----------------------------
        # -----------------------------
        # PAYMENT UPDATE
        # -----------------------------
        if loan_id:
            try:
                loan = Loan.objects.get(id=loan_id)

                def get_decimal(field_name):
                    try:
                        return Decimal(request.POST.get(field_name, "0") or "0")
                    except InvalidOperation:
                        return Decimal("0")

                cash  = get_decimal('cash')
                bank1 = get_decimal('bank1')
                bank2 = get_decimal('bank2')
                adj   = get_decimal('adj')

                total_payment = cash + bank1 + bank2 + adj
                original_payment = total_payment

                repayments = LoanRepayment.objects.filter(loan=loan).aggregate(
                    total_paid_principal=Coalesce(
                        Sum('paid_to_principal'), Value(0), output_field=DecimalField()
                    ),
                    total_paid_interest=Coalesce(
                        Sum('paid_to_interest'), Value(0), output_field=DecimalField()
                    )
                )

                total_paid_principal = Decimal(repayments['total_paid_principal'])
                total_paid_interest = Decimal(repayments['total_paid_interest'])

                remaining_interest = Decimal(loan.interest or 0) - total_paid_interest
                remaining_principal = Decimal(loan.amount or 0) - total_paid_principal

                # FIRST pay interest
                paid_to_interest = min(remaining_interest, total_payment)
                total_payment -= paid_to_interest

                # THEN pay principal
                paid_to_principal = min(remaining_principal, total_payment)
                total_payment -= paid_to_principal

                # -----------------------------
                # Handle repayment date
                # -----------------------------
                date_str = request.POST.get('date')
                selected_datetime = None
                if date_str:
                    try:
                        selected_datetime = datetime.strptime(date_str, "%Y-%m-%d")
                    except ValueError:
                        selected_datetime = None  # fallback to default

                # Save Repayment
                repayment = LoanRepayment.objects.create(
                    loan=loan,
                    total_payment=original_payment,
                    paid_to_interest=paid_to_interest,
                    paid_to_principal=paid_to_principal,
                    payment_mode='mixed',
                    remarks='',
                    cash=cash,
                    bank1=bank1,
                    bank2=bank2,
                    adj=adj,
                    code=generate_unique_repayment_code(),
                    type_of_loan=loan.type_of_loan
                )

                # Override created_at if a date was selected
                if selected_datetime:
                    repayment.created_at = selected_datetime
                    repayment.save(update_fields=['created_at'])

                # ----------------------------------------
                # AUTO CLOSE AFTER PAYMENT IF CLEARED
                # ----------------------------------------
                if (remaining_principal - paid_to_principal) <= 0 and (remaining_interest - paid_to_interest) <= 0:
                    loan.loan_status = "Closed"
                    loan.save(update_fields=["loan_status"])

                return JsonResponse({'status': 'success'})

            except Loan.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': 'Loan not found'}, status=404)
            except Exception as e:
                return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

    # -----------------------------
    # Ensure empty deposit types
    # -----------------------------
    existing_types = {x['loan_type'] for x in loans}

    for dep_type in deposit_types:
        if dep_type not in existing_types:
            loans.append({
                'loan_type': dep_type,
                'loan': None,
                'balance': Decimal('0'),
                'interest': Decimal('0'),
                'loan_code': '',
                
            })

    return render(request, 'deposits.html', {
        'loans': loans,
        'loan_types': deposit_types,
        'user_name': user_name,
        'gen_no': gen_no,
    })

from decimal import Decimal, InvalidOperation
from django.shortcuts import render
from django.http import JsonResponse
from django.db.models import Sum, Value, DecimalField
from django.db.models.functions import Coalesce
from .models import User, Loan, LoanRepayment, generate_unique_repayment_code
def others_view(request):
    other_types = ['ADMISSION FEES', 'OTHER RECEIPTS', 'CASH WITHDRAWALS']
    loans = []
    user_name = None
    gen_no = None

    if request.method == 'POST':
        gen_no = request.POST.get('gen_no', '').strip()
        loan_id = request.POST.get('loan_id')
        new_amount = Decimal(request.POST.get('new_amount', 0))
        loan_type = request.POST.get('loan_type')

        # ---------------- SEARCH / CREATE ----------------
        if gen_no and not loan_id:
            user = User.objects.filter(code=gen_no).first()
            user_name = user.name if user else 'Unknown'

            for oth_type in other_types:
                active_loan = Loan.objects.filter(
                    gen_no=gen_no,
                    type_of_loan=oth_type,
                    loan_status='Active'
                ).first()

                if new_amount > 0 and loan_type == oth_type:
                    if active_loan:
                        active_loan.amount += new_amount
                        active_loan.save(update_fields=['amount'])
                    else:
                        active_loan = Loan.objects.create(
                            gen_no=gen_no,
                            type_of_loan=oth_type,
                            amount=new_amount, # optional, if user provided a date
                            loan_status='Active'
                        )

                if active_loan:
                    total_paid = LoanRepayment.objects.filter(loan=active_loan).aggregate(
                        total=Coalesce(Sum('paid_to_principal'), Value(0), output_field=DecimalField())
                    )['total']

                    loans.append({
                        'loan_type': oth_type,
                        'loan': active_loan,
                        'balance': active_loan.amount - total_paid,
                        'loan_code': active_loan.id,
                    })

        # ---------------- STACK PAYMENT / RECEIPT ----------------
        if loan_id:
            loan = Loan.objects.get(id=loan_id)

            def d(name):
                try:
                    return Decimal(request.POST.get(name, '0') or '0')
                except:
                    return Decimal('0')

            cash, bank1, bank2, adj = d('cash'), d('bank1'), d('bank2'), d('adj')
            total = cash + bank1 + bank2 + adj

            total_paid = LoanRepayment.objects.filter(loan=loan).aggregate(
                total=Coalesce(Sum('paid_to_principal'), Value(0), output_field=DecimalField())
            )['total']

            remaining = loan.amount - total_paid
            pay_now = min(remaining, total)

            LoanRepayment.objects.create(
                loan=loan,
                total_payment=total,
                paid_to_principal=pay_now,
                paid_to_interest=Decimal('0'),
                payment_mode='mixed',
                cash=cash,
                bank1=bank1,
                bank2=bank2,
                adj=adj,
                code=generate_unique_repayment_code(),
                type_of_loan=loan.type_of_loan
            )

            if (remaining - pay_now) <= 0:
                loan.loan_status = 'Closed'
                loan.save(update_fields=['loan_status'])

            return JsonResponse({'status': 'success'})

    # ---------------- ENSURE EMPTY TYPES ----------------
    existing = {x['loan_type'] for x in loans}
    for t in other_types:
        if t not in existing:
            loans.append({
                'loan_type': t,
                'loan': None,
                'balance': Decimal('0'),
                'loan_code': ''
            })

    return render(request, 'others.html', {
        'loans': loans,
        'loan_types': other_types,
        'user_name': user_name,
        'gen_no': gen_no
    })

#--------------------------------------------------------------------------------------------------------------------------------------------------------------------------

def update_payment(request):
    if request.method == 'POST':
        loan_id = request.POST.get('loan_id')
        cash = Decimal(request.POST.get('cash', '0'))
        bank1 = Decimal(request.POST.get('bank1', '0'))
        bank2 = Decimal(request.POST.get('bank2', '0'))
        adj = Decimal(request.POST.get('adj', '0'))

        try:
            loan = Loan.objects.get(id=loan_id)
            total_payment = cash + bank1 + bank2 + adj

            # Find associated interest loan
            interest_type = f"{loan.type_of_loan.split()[0]} INTEREST"
            interest_loan = Loan.objects.filter(gen_no=loan.gen_no, type_of_loan=interest_type, loan_status='Active').first()

            # Deduct from interest loan first
            if interest_loan and interest_loan.amount > 0:
                if total_payment >= interest_loan.amount:
                    total_payment -= interest_loan.amount
                    interest_loan.amount = 0
                    interest_loan.loan_status = 'Closed'
                else:
                    interest_loan.amount -= total_payment
                    total_payment = 0
                interest_loan.save()

            # Deduct remaining from principal loan
            if total_payment > 0:
                if total_payment >= loan.amount:
                    total_payment -= loan.amount
                    loan.amount = 0
                    loan.loan_status = 'Closed'
                else:
                    loan.amount -= total_payment
                    total_payment = 0
                loan.save()

            # Record transaction
            LoanTransactions.objects.create(
                loan=loan,
                cash=cash,
                bank1=bank1,
                bank2=bank2,
                adj=adj
            )

            return JsonResponse({'status': 'success'})

        except Loan.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Loan not found.'}, status=404)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

    return JsonResponse({'status': 'error', 'message': 'Invalid request method.'}, status=405)


#--------------------------------------------------------------------------------------------------------------------------------------------------------------------------





def mtl_collection_view(request, loan_id):
    try:
        loan = Loan.objects.get(id=loan_id)
        gen_no = loan.gen_no
    except Loan.DoesNotExist:
        loan = get_object_or_404(InterestLoan, id=loan_id)
        related_loan = Loan.objects.filter(id=loan.original_loan_id).first()
        gen_no = related_loan.gen_no if related_loan else None

    if not gen_no:
        return render(request, 'mtl_collection.html', {
            'loan_data': [],
            'user_name': 'Unknown',
            'gen_no': 'Unknown',
        })

    user = User.objects.filter(code=gen_no).first()
    user_name = user.name if user else 'Unknown'
    loans = Loan.objects.filter(gen_no=gen_no)
    loan_data = []

    for loan in loans:
        repayments = LoanRepayment.objects.filter(loan=loan).annotate(
            month=TruncMonth('created_at')
        ).values('month').annotate(
            total_payment=Sum('total_payment'),
            paid_interest=Sum('paid_to_interest'),
            paid_principal=Sum('paid_to_principal')
        ).order_by('month')

        repayments_by_month = {
            r['month'].date(): {
                'total_payment': r['total_payment'] or Decimal(0),
                'paid_interest': r['paid_interest'] or Decimal(0),
                'paid_principal': r['paid_principal'] or Decimal(0)
            } for r in repayments
        }

        monthly_data = []
        # Use actual loan amount from DB as opening principal
        principal_opening = loan.amount  
        
        # Use actual loan.interest from DB as opening interest
        interest_opening = loan.interest  

        # Calculate month range based on repayments (or at least one month)
        if repayments_by_month:
            start_month = min(repayments_by_month.keys())
            end_month = max(repayments_by_month.keys())
        else:
            start_month = loan.created_at.date().replace(day=1)
            end_month = start_month

        current_month = start_month
        while current_month <= end_month:
            rep = repayments_by_month.get(current_month, {
                'total_payment': Decimal(0),
                'paid_interest': Decimal(0),
                'paid_principal': Decimal(0)
            })

            paid_principal = rep['paid_principal']
            paid_interest = rep['paid_interest']
            total_payment = rep['total_payment']

            # Calculate remaining balances ensuring no negative values
            principal_closing = max(Decimal(0), principal_opening - paid_principal)
            interest_closing = max(Decimal(0), interest_opening - paid_interest)

            monthly_data.append({
                'month': current_month,
                'total_collected': total_payment,
                'opening_balance': principal_opening,
                'opening_interest': interest_opening,
                'paid_principal': paid_principal,
                'paid_interest': paid_interest,
                'remaining_principal': principal_closing,
                'remaining_interest': interest_closing,
            })

            principal_opening = principal_closing
            interest_opening = interest_closing
            current_month += relativedelta(months=1)

            # Automatically close loan if principal is zero or less
            if principal_closing <= 0 and loan.loan_status != 'Closed':
                loan.loan_status = 'Closed'
                loan.save()

        loan_data.append({
            'loan': loan,
            'gen_no': gen_no,
            'transactions': monthly_data,
        })

    return render(request, 'mtl_collection.html', {
        'loan_data': loan_data,
        'user_name': user_name,
        'gen_no': gen_no,
    })

#--------------------------------------------------------------------------------------------------------------------------------------------------------------------------
def login_view(request):

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        if username == 'admin123' and password == 'admin123':
            user = authenticate(request, username=username, password=password)
            if user is not None:
               login(request,user)
            return redirect('loans')
        else:
            return render(request,'login.html',{'error':'Wrong Username or password'})
    else:
        return render(request, 'login.html')
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------------
from django.shortcuts import render, redirect
from django.contrib import messages
from django.core.mail import send_mail
from .forms import UserForm
from .models import User
import logging

# logger = logging.getLogger(__name__)
# def add_user(request):
#     if request.method == 'POST':
#         form = UserForm(request.POST)
#         if form.is_valid():
#             email = form.cleaned_data.get('Email')
            
#             # Check if email already exists
#             if User.objects.filter(Email=email).exists():
#                 msg = f"Email '{email}' already exists."
#                 print(msg)        # Print to console
#                 logger.warning(msg)
#                 messages.error(request, msg)
#                 return render(request, 'adduser.html', {'form': form})
            
#             try:
#                 # Save user
#                 user_instance = form.save()
                
#                 # Assuming code is generated by SQL trigger
#                 code = user_instance.code
                
#                 # Send email
#                 #subject = 'YOU HAVE REGISTERED WITH LOAN HUB'
#                 #message = f'Your UNIQUE ID is: {code}'
#                 #send_mail(subject, message, settings.DEFAULT_FROM_EMAIL, [email], fail_silently=False)
#                 #success_msg = f'User added successfully! Email sent to {email}'
#                 #print(success_msg)  # Print to console
#                 #logger.info(success_msg)
#                 #messages.success(request, success_msg)
#                 print("this is the message for testing",file=sys.stderr)
#                 return redirect('loans')
            
#             except User.DoesNotExist as e:
#                 print("User retrieval failed:", e)
#                 logger.error("User retrieval failed", exc_info=True)
#                 messages.error(request, 'Failed to find User with the provided Email.')
            
#             except Exception as e:
#                 print("An unexpected error occurred:", e)
#                 logger.error("Unexpected error in add_user view", exc_info=True)
#                 messages.error(request, f'An error occurred: {str(e)}')
        
#         else:
#             print("Form validation failed:", form.errors)
#             logger.warning("Form validation failed: %s", form.errors)
#             messages.error(request, 'Failed to add User. Please correct the errors below.')
    
#     else:
#         form = UserForm()
    
#     return render(request, 'adduser.html', {'form': form})
def add_user(request):
    if request.method == 'POST':
        form = UserForm(request.POST)
        if form.is_valid():
            try:
                user_instance = form.save()
                messages.success(request, "User added successfully!")
                return redirect('loans')

            except Exception as e:
                logger.error("Unexpected error in add_user view", exc_info=True)
                messages.error(request, f"An error occurred: {str(e)}")
        else:
            print("Form validation failed:", form.errors)
            messages.error(request, "Please correct the errors below.")
    else:
        form = UserForm()

    return render(request, 'adduser.html', {'form': form})
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------------

def send_loan_email(receiver_email):
    subject = 'Loan Added'
    message = 'A loan has been added for your account.'
    from_email = 'loanhub@arshithtech.in'  # Update with your email
    send_mail(subject, message, from_email, [receiver_email])
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------------
from decimal import Decimal
from django.shortcuts import render, redirect
from .models import Loan, User
from .forms import LoanForm


from datetime import datetime

from datetime import datetime

def add_loan(request):
    if request.method == "POST":
        form = LoanForm(request.POST)
        if form.is_valid():
            gen_no = form.cleaned_data["gen_no"]
            loan_date_str = request.POST.get("loan_date")

            # Validate user exists
            if not User.objects.filter(code=gen_no).exists():
                form.add_error("gen_no", "User with this Gen.no does not exist.")
            else:
                # Create object but don't save yet
                loan = form.save(commit=False)

                # Override created_at if user selected date
                if loan_date_str:
                    try:
                        loan.created_at = datetime.strptime(loan_date_str, "%Y-%m-%d")
                    except ValueError:
                        pass

                loan.save()
                return redirect("/loans")
    else:
        form = LoanForm()

    return render(request, "addloan.html", {"form": form})

#--------------------------------------------------------------------------------------------------------------------------------------------------------------------------

def interest_rate_view(request):
    interest_rates = InterestRate.objects.all()
    context = {
        'interest_rates': interest_rates
    }
    return render(request, 'interestrate.html',context)
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------------

def Update_intrest_rate(request):
    if request.method == 'POST':
        form = IntrestForm(request.POST)
        if form.is_valid():
            type_of_receipt = form.cleaned_data['Type_of_Receipt']
            new_interest_rate = form.cleaned_data['interest']
            # Update the interest rate in the database
            interest_rate = InterestRate.objects.get(Type_of_Receipt=type_of_receipt)
            interest_rate.interest = new_interest_rate
            interest_rate.save()
            return redirect('intrest_rate')  # Redirect to the interest rate view page
    else:
        form = IntrestForm()
    return render(request, 'interestrate.html',{'form':form})

#--------------------------------------------------------------------------------------------------------------------------------------------------------------------------

def closed_loans_view(request):
    closed_loans = []
    user_name = None
    gen_no = request.GET.get("gen_no") or request.POST.get("gen_no")  # works for GET + POST

    # Show all closed loans if no filter
    if not gen_no:
        closed_loans = Loan.objects.filter(loan_status="Closed")
    else:
        # filter loans
        closed_loans = Loan.objects.filter(gen_no__icontains=gen_no, loan_status="Closed")

        # fetch user for display UI
        user = User.objects.filter(code__icontains=gen_no).first()
        if user:
            user_name = user.name

    return render(request, 'closed_loans.html', {
        'closed_loans': closed_loans,
        'user_name': user_name,
        'gen_no': gen_no,
    })


#--------------------------------------------------------------------------------------------------------------------------------------------------------------------------
from decimal import Decimal, InvalidOperation
from django.shortcuts import render
from .models import Loan, LoanRepayment

SWAP_LOAN_TYPES = ['FIXED DEPOSITS', 'THRIFT FUNDS', 'WELFARE COLLECTIONS']


def safe_decimal(value):
    """Convert a value to Decimal safely, treating invalid strings as 0."""
    try:
        return Decimal(value)
    except (InvalidOperation, TypeError, ValueError):
        return Decimal('0')
from django.shortcuts import render
from django.db.models import F
from decimal import Decimal

def safe_decimal(value):
    try:
        return Decimal(value or 0)
    except:
        return Decimal(0)

from decimal import Decimal
from datetime import datetime

def cash_book(request):
    # ------------------ Fetch existing transactions ------------------
    transactions = list(
        LoanRepayment.objects.select_related('loan').order_by('-created_at', '-id')
    )

    payments = list(
        Loan.objects.all().order_by('-created_at', '-id')
    )

    from decimal import Decimal

    from decimal import Decimal

    # ------------------ Include Cash Entries ------------------
    cash_entries = CashEntry.objects.all().order_by('-datetime')

    class CashBookEntry:
        def __init__(self, cash_entry):
            raw_amount = Decimal(cash_entry.amount)
            self.amount = abs(raw_amount)

            self.cash = self.amount if cash_entry.type_of_cash == "Cash" else Decimal('0')
            self.bank1 = self.amount if cash_entry.type_of_cash == "Bank1" else Decimal('0')
            self.bank2 = self.amount if cash_entry.type_of_cash == "Bank2" else Decimal('0')
            self.adj = self.amount if cash_entry.type_of_cash == "Adjustment" else Decimal('0')

            # 🔥 ADD THIS LINE
            self.type_of_loan = cash_entry.type_of_loan

            self.loan = type('LoanObj', (object,), {
                'id': cash_entry.code,
                'type_of_loan': cash_entry.type_of_loan
            })()

            self.remarks = cash_entry.remarks
            self.code = cash_entry.code
            self.created_at = cash_entry.datetime
            self.is_negative = raw_amount < 0


    # Append to transactions or payments
    for entry in cash_entries:
        cb_entry = CashBookEntry(entry)

        # ✅ Negative values always go to payments
        if cb_entry.is_negative:
            payments.append(cb_entry)

        # Existing swap logic
        elif entry.type_of_cash in SWAP_LOAN_TYPES:
            payments.append(cb_entry)

        else:
            transactions.append(cb_entry)


    # ---------------- Swap Logic ----------------
    new_transactions, new_payments = [], []

    for t in transactions:
        loan_type = getattr(t.loan, 'type_of_loan', None)
        if loan_type in SWAP_LOAN_TYPES:
            new_payments.append(t)
        else:
            new_transactions.append(t)

    for p in payments:
        # Determine loan type depending on object
        if hasattr(p, 'loan'):
            loan_type = getattr(p.loan, 'type_of_loan', None)  # LoanRepayment
        else:
            loan_type = getattr(p, 'type_of_loan', None)       # Loan or CashBookEntry

        if loan_type in SWAP_LOAN_TYPES:
            new_transactions.append(p)
        else:
            new_payments.append(p)

    transactions = new_transactions
    payments = new_payments

    # ---------------- Totals ----------------
    def safe_decimal(value):
        try:
            return Decimal(value or 0)
        except:
            return Decimal(0)
# ---------------- Include OtherCashTransaction BEFORE totals ----------------

    other_receipts = list(
        OtherCashTransaction.objects.filter(transaction_type='RECEIPT')
    )

    other_payments = list(
        OtherCashTransaction.objects.filter(transaction_type='PAYMENT')
    )

    transactions.extend(other_receipts)
    payments.extend(other_payments)

    # ---------------- Sort by time ----------------
    transactions.sort(key=lambda x: x.created_at, reverse=True)
    payments.sort(key=lambda x: x.created_at, reverse=True)

    # ---------------- Totals (NO adj field) ----------------
    def get_adj(obj):
        return safe_decimal(obj.adj) if hasattr(obj, 'adj') else Decimal('0')


    # ---------------- Totals ----------------
    transactions_totals = {
        'total_cash': sum(safe_decimal(getattr(t, 'cash', 0)) for t in transactions),
        'total_bank1': sum(safe_decimal(getattr(t, 'bank1', 0)) for t in transactions),
        'total_bank2': sum(safe_decimal(getattr(t, 'bank2', 0)) for t in transactions),
        'total_adj': sum(get_adj(t) for t in transactions),
    }

    payments_totals = {
        'total_cash': sum(safe_decimal(getattr(p, 'cash', 0)) for p in payments),
        'total_bank1': sum(safe_decimal(getattr(p, 'bank1', 0)) for p in payments),
        'total_bank2': sum(safe_decimal(getattr(p, 'bank2', 0)) for p in payments),
        'total_adj': sum(get_adj(p) for p in payments),
        'total_amount': sum(
            safe_decimal(getattr(p, 'amount', 0))
            for p in payments
        )
    }


    # ================== CARD TOTALS (FOR CASH OVERVIEW CARDS ONLY) ==================

    receipts_totals = {
        'cash': sum(safe_decimal(getattr(t, 'cash', 0)) for t in transactions),
        'bank1': sum(safe_decimal(getattr(t, 'bank1', 0)) for t in transactions),
        'bank2': sum(safe_decimal(getattr(t, 'bank2', 0)) for t in transactions),
        'adj': sum(get_adj(t) for t in transactions),
    }

    payments_totals_cards = {
        'cash': sum(safe_decimal(getattr(p, 'cash', 0)) for p in payments),
        'bank1': sum(safe_decimal(getattr(p, 'bank1', 0)) for p in payments),
        'bank2': sum(safe_decimal(getattr(p, 'bank2', 0)) for p in payments),
        'adj': sum(get_adj(p) for p in payments),
    }

    withdrawals = {
        'cash': receipts_totals['cash'] - payments_totals_cards['cash'],
        'bank1': receipts_totals['bank1'] - payments_totals_cards['bank1'],
        'bank2': receipts_totals['bank2'] - payments_totals_cards['bank2'],
        'adj': receipts_totals['adj'] - payments_totals_cards['adj'],
    }

    # ---------------- Final balances ----------------
    total_receipts = (
        transactions_totals['total_cash'] +
        transactions_totals['total_bank1'] +
        transactions_totals['total_bank2'] +
        transactions_totals['total_adj']
    )

    total_payments = payments_totals['total_amount']
    available_balance = total_receipts - total_payments


    return render(request, 'cash_book.html', {
        'transactions': transactions,
        'payments': payments,
        'transactions_totals': transactions_totals,
        'payments_totals': payments_totals,
        'available_balance': available_balance,

        # 🔥 NEW (for cards)
        'withdrawals': withdrawals,
    })



#------------------------------------------------------------------------------------------------------------------------------------------------------------------------

from datetime import datetime



def loan_transactions_detail(request, loan_id, month):
    # Get the loan object (from either Loan or InterestLoan)
    try:
        loan = Loan.objects.get(id=loan_id)
    except Loan.DoesNotExist:
        loan = get_object_or_404(InterestLoan, id=loan_id)

    # Parse the month
    selected_month = month + '-01'
    try:
        month_date = datetime.strptime(selected_month, '%Y-%m-%d')
        month_name = calendar.month_name[month_date.month]
        year = month_date.year
    except ValueError:
        return HttpResponseBadRequest("Invalid month format.")

    # Get repayments for the selected loan and month
    repayments = LoanRepayment.objects.filter(
        loan=loan,
        created_at__year=month[:4],
        created_at__month=month[5:]
    )

    if request.method == 'POST':
        for repayment in repayments:
            save_button = f"save_{repayment.code}"
            delete_button = f"delete_{repayment.code}"

            # Handle deletion
            if delete_button in request.POST:
                repayment.delete()
                return redirect('loan_transactions_detail', loan_id=loan_id, month=month)

            # Handle update/save
            if save_button in request.POST:
                def get_decimal(field_name, default):
                    try:
                        return Decimal(request.POST.get(field_name, default).strip() or default)
                    except (InvalidOperation, TypeError):
                        return Decimal(default)

                cash = get_decimal(f"cash_{repayment.code}", '0')
                bank1 = get_decimal(f"bank1_{repayment.code}", '0')
                bank2 = get_decimal(f"bank2_{repayment.code}", '0')
                adj = get_decimal(f"adj_{repayment.code}", '0')
                total_payment = cash + bank1 + bank2 + adj

                # Calculate remaining principal and interest
                previous = LoanRepayment.objects.filter(
                    loan=loan
                ).exclude(id=repayment.id).aggregate(
                    total_paid_principal=Sum('paid_to_principal'),
                    total_paid_interest=Sum('paid_to_interest')
                )

                total_paid_principal = previous['total_paid_principal'] or Decimal('0')
                total_paid_interest = previous['total_paid_interest'] or Decimal('0')

                loan_amount = loan.amount or Decimal('0')
                loan_interest = loan.interest or Decimal('0')

                remaining_principal = loan_amount - total_paid_principal
                remaining_interest = loan_interest - total_paid_interest

                paid_to_interest = min(remaining_interest, total_payment)
                leftover = total_payment - paid_to_interest
                paid_to_principal = min(remaining_principal, leftover)

                # Save updated repayment
                repayment.cash = cash
                repayment.bank1 = bank1
                repayment.bank2 = bank2
                repayment.adj = adj
                repayment.total_payment = total_payment
                repayment.paid_to_interest = paid_to_interest
                repayment.paid_to_principal = paid_to_principal
                repayment.save()

                return redirect('loan_transactions_detail', loan_id=loan_id, month=month)

    # Prepare transaction data for the template
    transaction_data = []
    for r in repayments:
        transaction_data.append({
            'total_payment': r.total_payment,
            'paid_to_interest': r.paid_to_interest,
            'paid_to_principal': r.paid_to_principal,
            'payment_mode': r.payment_mode,
            'remarks': r.remarks,
            'created_at': r.created_at,
            'cash': r.cash,
            'bank1': r.bank1,
            'bank2': r.bank2,
            'adj': r.adj,
            'code': r.code,
        })

    return render(request, 'loan_transactions_detail.html', {
        'loan': loan,
        'transactions': transaction_data,
        'month': f"{month_name} {year}",
    })


#------------------------------------------------------------------------------------------------------------------------------------------------------------------------
from datetime import datetime

def submit_new_table(request):
    if request.method == 'POST':
        gen_no = request.POST.get('gen_no')
        loan_type = request.POST.get('Loan Type')
        amount = request.POST.get('Amount')
        cash = request.POST.get('Cash')
        online = request.POST.get('Online')
        bank1 = request.POST.get('Bank1')
        bank2 = request.POST.get('Bank2')
        adj = request.POST.get('Adj')
        loan_date_str = request.POST.get('date')   # 🔥 FIX HERE

        if not gen_no or not loan_type or not amount:
            return JsonResponse({'status': 'error', 'message': 'Required fields are missing.'}, status=400)

        user = User.objects.filter(code=gen_no).first()
        if not user:
            return JsonResponse({'status': 'error', 'message': 'User with this Gen.no does not exist.'}, status=404)

        loan_data = {
            'gen_no': gen_no,
            'name': user.name,
            'amount': amount,
            'cash': cash,
            'online': online,
            'bank1': bank1,
            'bank2': bank2,
            'adj': adj,
            'type_of_loan': loan_type
        }

        form = LoanForm(loan_data)

        if form.is_valid():
            loan = form.save(commit=False)

            # ---------------- FIX DATE HANDLING ----------------
            if loan_date_str:
                try:
                    loan.created_at = datetime.strptime(loan_date_str, "%Y-%m-%d")
                except ValueError:
                    pass
            # --------------------------------------------------

            loan.interest = 0
            loan.save()

            send_loan_email(user.Email)

            return JsonResponse({'status': 'success', 'message': 'Loan added successfully!'}, status=200)

        return JsonResponse({'status': 'error', 'message': 'Form validation failed.'}, status=400)

    return JsonResponse({'status': 'error', 'message': 'Invalid request method.'}, status=405)

#--------------------------------------------------------------------------------------------------------------------------------------------------------------------------
def search_user_codes(request):
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        term = request.GET.get('term', '')
        if term:
            # Normalize the input by removing spaces and converting to lower case
            normalized_term = term.replace(" ", "").lower()

            # Search in User model for code
            users = User.objects.filter(code__icontains=normalized_term)
            if users.exists():
                user_code_name = list(users.values('code', 'name'))
                return JsonResponse(user_code_name, safe=False)  # Return user code and name

            # Search in Loan model for code
            loans = Loan.objects.filter(code__icontains=normalized_term)
            if loans.exists():
                # If loan is found, get the gen_no and find users associated with that gen_no
                gen_no = list(loans.values_list('gen_no', flat=True))
                
                # Find users by the found gen_no
                users = User.objects.filter(code__in=gen_no)  # Adjust field as necessary
                if users.exists():
                    user_code_name = list(users.values('code', 'name'))
                    return JsonResponse(user_code_name, safe=False)  # Return user code and name

                return JsonResponse([], safe=False)
            # Search in User model for name with normalized check
            users_by_name = User.objects.filter(
                Q(name__icontains=normalized_term)
            )
            if users_by_name.exists():
                user_code_name_by_name = list(users_by_name.values('code', 'name'))
                return JsonResponse(user_code_name_by_name, safe=False)  # Return user code and name by name
            
            # Search InterestLoan for loan codes associated with original loans
            interest_loans = InterestLoan.objects.filter(code__icontains=normalized_term)
            if interest_loans.exists():
                # Assuming `original_loan` is a ForeignKey to Loan
                loan_ids = interest_loans.values_list('original_loan', flat=True)
                loans_with_gen_no = Loan.objects.filter(id__in=loan_ids).values_list('gen_no', flat=True)
                
                # Find users by the found gen_no
                users = User.objects.filter(code__in=list(loans_with_gen_no))  # Adjust field as necessary
                if users.exists():
                    user_code_name = list(users.values('code', 'name'))
                    return JsonResponse(user_code_name, safe=False)  # Return user code and name

                return JsonResponse([], safe=False)

        return JsonResponse([], safe=False)  # Return an empty list if no matches are found
    return JsonResponse({'error': 'Invalid request'}, status=400)

#------------------------------------------------------------------------------------------------------------------------------------------------------------------------


from decimal import Decimal, ROUND_HALF_UP, ROUND_DOWN
from django.db.models import F
from .models import Loan, InterestRate

EXCLUDED_TYPES = ['ADMISSION FEES', 'OTHER RECEIPTS', 'CASH WITHDRAWALS']
def update_loans_interest():
    print("DAILY INTEREST JOB STARTED")

    loans = Loan.objects.filter(
        loan_status='Active'
    ).exclude(type_of_loan__in=EXCLUDED_TYPES)

    print("Loans found:", loans.count())

    for loan in loans:
        loan_amount = loan.amount
        if not loan_amount:
            continue

        try:
            rate_obj = InterestRate.objects.get(
                Type_of_Receipt=loan.type_of_loan
            )
            annual_rate = Decimal(rate_obj.interest) / Decimal('100')
        except InterestRate.DoesNotExist:
            print("Rate missing for:", loan.type_of_loan)
            continue

        daily_interest = loan_amount * annual_rate / Decimal('365')

        if daily_interest % 1 == Decimal('0.50'):
            daily_interest = daily_interest.quantize(
                Decimal('1'), rounding=ROUND_DOWN
            )
        else:
            daily_interest = daily_interest.quantize(
                Decimal('1'), rounding=ROUND_HALF_UP
            )

        print(f"Loan {loan.id} → interest added = {daily_interest}")

        Loan.objects.filter(pk=loan.pk).update(
            interest=F('interest') + daily_interest
        )

    print("DAILY INTEREST JOB FINISHED")

# Start the thread to update interest loans
# threading.Thread(target=update_loans_interest, daemon=True).start() 
# loan_hub/views.py
import time
from decimal import Decimal
from threading import Thread
from django.db import connection
from django.db.models import F
from .models import Loan

EXCLUDED_TYPES = ['ADMISSION FEES', 'OTHER RECEIPTS', 'CASH WITHDRAWALS']

def table_exists(table_name):
    """Check if table exists in the database"""
    return table_name in connection.introspection.table_names()

# def update_loans_interest():
#     while True:
#         pass
# Home view to display InterestLoan data
def home(request):
    entries = InterestLoan.objects.all()
    return render(request, 'home.html', {'entries': entries})
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import AddCash

def add_cash_view(request):
    if request.method == 'POST':
        amount = request.POST.get('amount')
        type_of_cash = request.POST.get('type_of_cash')
        remarks = request.POST.get('remarks', '')

        # validation
        if not amount or not type_of_cash:
            messages.error(request, "Amount and Type are required.")
            return redirect('add_cash')

        try:
            amount = float(amount)
        except ValueError:
            messages.error(request, "Amount must be a valid number.")
            return redirect('add_cash')

        # save
        AddCash.objects.create(
            amount=amount,
            type_of_cash=type_of_cash,
            remarks=remarks
        )

        messages.success(request, "Cash entry added successfully!")
        return redirect('add_cash')

    return render(request, 'add_cash.html')

from django.shortcuts import render, redirect
from django.contrib import messages
from .models import AddCash

def add_cash_view(request):
    if request.method == 'POST':
        amount = request.POST.get('amount')
        type_of_cash = request.POST.get('type_of_cash')
        remarks = request.POST.get('remarks', '')

        # validation
        if not amount or not type_of_cash:
            messages.error(request, "Amount and Type are required.")
            return redirect('add_cash')

        try:
            amount = float(amount)
        except ValueError:
            messages.error(request, "Amount must be a valid number.")
            return redirect('add_cash')

        # save
        AddCash.objects.create(
            amount=amount,
            type_of_cash=type_of_cash,
            remarks=remarks
        )

        messages.success(request, "Cash entry added successfully!")
        return redirect('add_cash')

    # fetch all entries
    cash_entries = AddCash.objects.all().order_by('-datetime')  # latest first
    return render(request, 'add_cash.html', {"cash_entries": cash_entries})






def reports_view(request):
    # Your logic to fetch report data
    users = User.objects.all().order_by('name')  # Fetch all users
    context = {}
    return render(request, 'reports.html', context)

import csv

from decimal import Decimal
from django.shortcuts import render
from django.http import HttpResponse
from django.utils.timezone import localtime
from openpyxl import Workbook
from .models import User, Loan, LoanRepayment

# Categories
LOAN_REPAYMENTS = ['MTL LOAN', 'FDL LOAN', 'KVP/NSC LOAN']
DEPOSIT_REPAYMENTS = ['FIXED DEPOSITS', 'THRIFT FUNDS', 'WELFARE COLLECTIONS']
OTHER_RECEIPTS = ['ADMISSION FEES', 'OTHER RECEIPTS', 'CASH WITHDRAWALS']

def safe_decimal(value):
    try:
        return Decimal(value or 0)
    except:
        return Decimal(0)

# ---------------- REPORTS PAGE ----------------
def reports_view(request):
    users = User.objects.all().order_by('name')
    return render(request, 'reports.html', {'users': users})

# ---------------- DOWNLOAD USERS ----------------
from datetime import datetime
from django.http import HttpResponse
from openpyxl import Workbook
from .models import User

def download_users(request):
    users = User.objects.all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Users"

    ws.append(["ID", "Name", "Mobile", "Email", "Code", "Age", "Bank Account 1", "IFSC 1", "Bank Account 2", "IFSC 2"])

    for u in users:
        ws.append([
            u.id, u.name, u.Mobile, u.Email, 
            u.code, u.Age, 
            u.AccountNo1, u.IFSCcode, 
            u.AccountNo2, u.IFSCcode2
        ])

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=users_{timestamp}.xlsx'

    wb.save(response)
    return response


# ---------------- DOWNLOAD GENERAL RECEIPTS ----------------
from datetime import datetime
from django.http import HttpResponse
from openpyxl import Workbook
from django.utils.timezone import localtime

def download_receipts(request):
    records = Loan.objects.filter(
        type_of_loan__in=LOAN_REPAYMENTS + DEPOSIT_REPAYMENTS + OTHER_RECEIPTS
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Receipts"

    ws.append(["ID", "User", "Type", "Cash", "Bank1", "Bank2", "Adjustment", "Amount", "Date"])

    for r in records:
        date_val = localtime(r.created_at).replace(tzinfo=None)
        ws.append([
            r.id,
            r.name,
            r.type_of_loan,
            safe_decimal(r.cash),
            safe_decimal(r.bank1),
            safe_decimal(r.bank2),
            safe_decimal(r.adj),
            safe_decimal(r.amount),
            date_val
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    # FIXED HERE ↓↓↓
    response['Content-Disposition'] = (
        f'attachment; filename=payments_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )

    wb.save(response)
    return response

# ---------------- DOWNLOAD GENERAL PAYMENTS ----------------
from datetime import datetime
from django.http import HttpResponse
from openpyxl import Workbook
from django.utils.timezone import localtime

def download_payments():
    records = LoanRepayment.objects.all().select_related('loan')

    wb = Workbook()
    ws = wb.active
    ws.title = "Payments"

    ws.append([
        "ID", "User", "Type", "Cash", "Bank1", "Bank2",
        "Adjustment", "Total Payment", "Date"
    ])

    for r in records:
        date_val = localtime(r.created_at).replace(tzinfo=None)
        ws.append([
            r.id,
            r.loan.name,
            r.loan.type_of_loan,
            safe_decimal(r.cash),
            safe_decimal(r.bank1),
            safe_decimal(r.bank2),
            safe_decimal(r.adj),
            safe_decimal(r.total_payment),
            date_val
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    # FIXED HERE ↓↓↓
    response['Content-Disposition'] = (
        f'attachment; filename=reciepts_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )

    wb.save(response)
    return response

# ---------------- DOWNLOAD USER-SPECIFIC REPORT ----------------
from decimal import Decimal
from django.http import HttpResponse
from django.utils.timezone import localtime
from openpyxl import Workbook
from datetime import datetime  # ✅ Use this style
from .models import User, Loan, LoanRepayment

# Categories
LOAN_REPAYMENTS = ['MTL LOAN', 'FDL LOAN', 'KVP/NSC LOAN']
DEPOSIT_REPAYMENTS = ['FIXED DEPOSITS', 'THRIFT FUNDS', 'WELFARE COLLECTIONS']
OTHER_RECEIPTS = ['ADMISSION FEES', 'OTHER RECEIPTS', 'CASH WITHDRAWALS']

def safe_decimal(value):
    try:
        return Decimal(value or 0)
    except:
        return Decimal(0)

def download_user_report(request):
    if request.method != "POST":
        return HttpResponse("Invalid request")

    user_id = request.POST.get('user')
    category = request.POST.get('category')

    if not user_id or not category:
        return HttpResponse("User and category are required")

    # Get user object
    try:
        user_obj = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return HttpResponse("User not found")

    records = []

    # Filter records based on category
    if category == 'loan_repayments':
        records = LoanRepayment.objects.filter(
            loan__name=user_obj.name,
            loan__type_of_loan__in=LOAN_REPAYMENTS
        ).select_related('loan')

    elif category == 'deposit_repayments':
        records = LoanRepayment.objects.filter(
            loan__name=user_obj.name,
            loan__type_of_loan__in=DEPOSIT_REPAYMENTS
        ).select_related('loan')

    elif category == 'other_receipts':
        records = Loan.objects.filter(
            name=user_obj.name,
            type_of_loan__in=OTHER_RECEIPTS
        )

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "User Report"

    # Header
    ws.append([
        "ID", "User", "Type", "Cash", "Bank1", "Bank2", "Adjustment", "Amount/Total Payment", "Date"
    ])

    # Add data rows
    for r in records:
        date_val = localtime(r.created_at).replace(tzinfo=None)
        if hasattr(r, 'loan'):  # LoanRepayment
            ws.append([
                r.id,
                r.loan.name,
                r.loan.type_of_loan,
                safe_decimal(r.cash),
                safe_decimal(r.bank1),
                safe_decimal(r.bank2),
                safe_decimal(r.adj),
                safe_decimal(r.total_payment),
                date_val
            ])
        else:  # Loan (Other Receipts)
            ws.append([
                r.id,
                r.name,
                r.type_of_loan,
                safe_decimal(r.cash),
                safe_decimal(r.bank1),
                safe_decimal(r.bank2),
                safe_decimal(r.adj),
                safe_decimal(r.amount),
                date_val
            ])

    # Prepare HTTP response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"user_report_{user_obj.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'

    wb.save(response)
    return response

# donot touch above by TARUN AMARANENI

from decimal import Decimal
from django.db.models import Sum, Value, DecimalField
from django.db.models.functions import Coalesce
from .models import Loan, LoanRepayment, OtherCashTransaction, User


# Safe decimal conversion
def safe_decimal(value):
    try:
        return Decimal(value or 0)
    except Exception:
        return Decimal('0')


def reports_view(request):

    # ===================== CATEGORY SETUP =====================
    LOAN_TYPES = ['MTL LOAN', 'FDL LOAN', 'KVP/NSC LOAN']
    DEPOSIT_TYPES = ['FIXED DEPOSITS', 'THRIFT FUNDS', 'WELFARE COLLECTIONS']
    OTHER_TYPES = ['ADMISSION FEES', 'OTHER RECEIPTS', 'SALARY PAID','OFFICE EXPENSES','OTHER PAYMENTS']

    loans_qs = Loan.objects.filter(type_of_loan__in=LOAN_TYPES).order_by('gen_no', 'id')
    deposits_qs = Loan.objects.filter(type_of_loan__in=DEPOSIT_TYPES).order_by('gen_no', 'id')

    # ===================== COMMON FUNCTION =====================
    def compute_loan_fields(qs):
        rows = []
        for loan in qs:
            agg = LoanRepayment.objects.filter(loan=loan).aggregate(
                total_paid_principal=Coalesce(Sum('paid_to_principal'), Value(0), output_field=DecimalField()),
                total_paid_interest=Coalesce(Sum('paid_to_interest'), Value(0), output_field=DecimalField()),
            )

            paid_principal = safe_decimal(agg['total_paid_principal'])
            paid_interest = safe_decimal(agg['total_paid_interest'])

            amount = safe_decimal(loan.amount)
            interest = safe_decimal(loan.interest)

            remaining_principal = amount - paid_principal
            remaining_interest = interest - paid_interest

            if remaining_principal <= 0 and remaining_interest <= 0:
                continue

            rows.append({
                'id': loan.id,
                'gen_no': loan.gen_no,
                'name': loan.name,
                'ref': loan.code or '',
                'type': loan.type_of_loan,
                'amount': amount,
                'balance': remaining_principal,
                'interest': remaining_interest,
                'created_at': getattr(loan, 'created_at', None),
            })
        return rows

    loans = compute_loan_fields(loans_qs)
    deposits = compute_loan_fields(deposits_qs)

    # ===================== ✅ OTHERS (FIXED PROPERLY) =====================
# ===================== OTHERS (ALL RECORDS) =====================
    others = []

    transactions = OtherCashTransaction.objects.all().order_by('-id')

    # map users by code (fast lookup)
    user_map = {u.code: u.name for u in User.objects.all()}

    for txn in transactions:
        others.append({
            'id': txn.id,
            'gen_no': txn.gen_no,
            'name': user_map.get(txn.gen_no, 'Unknown'),
            'type_of_loan': txn.type_of_loan,   # ✅ RECEIPT / PAYMENT
            'ref': txn.code or '',           # ✅ code
            'amount': safe_decimal(txn.amount),
            'created_at': txn.created_at,
        })


    # ===================== RENDER =====================
    return render(request, 'reports.html', {
        'LOAN_TYPES': LOAN_TYPES,
        'DEPOSIT_TYPES': DEPOSIT_TYPES,
        'OTHER_TYPES': OTHER_TYPES,
        'loans': loans,
        'deposits': deposits,
        'others': others,
    })


from django.shortcuts import render
from .models import User

def download_reports(request):
    users = User.objects.all().order_by('name')  # fetch all users
    return render(request, 'download_reports.html', {'users': users})


from django.shortcuts import render, redirect
from django.contrib import messages
from .models import CashEntry
import random

def cash_entry_view(request):
    if request.method == "POST":
        amount = request.POST.get("amount")
        type_of_cash = request.POST.get("type_of_cash")
        remarks = request.POST.get("remarks", "")

        if amount and type_of_cash:
            # Auto-generate unique code starting with CA
            while True:
                code = f"CA{random.randint(1000,9999)}"
                if not CashEntry.objects.filter(code=code).exists():
                    break

            # Create CashEntry with type_of_loan = "AddCash"
            CashEntry.objects.create(
                amount=amount,
                type_of_cash=type_of_cash,
                remarks=remarks,
                code=code,
                type_of_loan="AddCash"  # <-- automatically set
            )

            messages.success(request, f"Amount added successfully! Code: {code}")
            return redirect("cash_entry")  # your URL name

        else:
            messages.error(request, "Please fill all required fields.")

    cash_entries = CashEntry.objects.all().order_by("-datetime")
    return render(request, "add_cash.html", {"cash_entries": cash_entries})


# today

from decimal import Decimal
from django.shortcuts import render
from django.db.models import Sum, Value, DecimalField
from django.db.models.functions import Coalesce
from .models import Loan, LoanRepayment


def safe_decimal(value):
    try:
        return Decimal(value or 0)
    except:
        return Decimal('0')


def download_reports_view(request):

    # CATEGORIES
    LOAN_TYPES = ['MTL LOAN', 'FDL LOAN', 'KVP/NSC LOAN']
    DEPOSIT_TYPES = ['FIXED DEPOSITS', 'THRIFT FUNDS', 'WELFARE COLLECTIONS']
    OTHER_TYPES = ['ADMISSION FEES', 'OTHER RECEIPTS', 'CASH WITHDRAWALS']

    # Get selected type from GET
    selected_type = request.GET.get("type")       # ex: ?type=MTL LOAN

    rows = []
    summary = {"interest": 0, "payments": 0, "deposits": 0}

    if selected_type:
        loans = Loan.objects.filter(type_of_loan=selected_type).order_by("gen_no")

        for loan in loans:
            agg = LoanRepayment.objects.filter(loan=loan).aggregate(
                paid_principal=Coalesce(Sum("paid_to_principal"), Value(0), output_field=DecimalField()),
                paid_interest=Coalesce(Sum("paid_to_interest"), Value(0), output_field=DecimalField()),
                deposits=Coalesce(Sum("deposits"), Value(0), output_field=DecimalField())
            )

            paid_principal = safe_decimal(agg["paid_principal"])
            paid_interest = safe_decimal(agg["paid_interest"])
            deposits = safe_decimal(agg["deposits"])

            interest_remaining = loan.interest - paid_interest
            principal_remaining = loan.amount - paid_principal

            rows.append({
                "loan": loan,
                "principal_remaining": principal_remaining,
                "interest_remaining": interest_remaining,
                "total_payments": paid_principal + paid_interest,
                "total_deposits": deposits
            })

            # Build summary cards
            summary["interest"] += interest_remaining
            summary["payments"] += paid_principal + paid_interest
            summary["deposits"] += deposits

    context = {
        "LOAN_TYPES": LOAN_TYPES,
        "DEPOSIT_TYPES": DEPOSIT_TYPES,
        "OTHER_TYPES": OTHER_TYPES,

        "selected_type": selected_type,
        "rows": rows,
        "summary": summary,
    }

    return render(request, "download_reports.html", context)


# today 2
# from decimal import Decimal
# from django.shortcuts import render
# from .models import Loan, LoanRepayment
# from django.db.models import Sum, Value, DecimalField
# from django.db.models.functions import Coalesce

# SWAP_LOAN_TYPES = ['FIXED DEPOSITS', 'THRIFT FUNDS', 'WELFARE COLLECTIONS']

# def safe_decimal(value):
#     try:
#         return Decimal(value)
#     except:
#         return Decimal('0')

# def get_loan_type(obj):
#     if hasattr(obj, "loan"):
#         return getattr(obj.loan, "type_of_loan", "-")
#     return getattr(obj, "type_of_loan", "-")

# def reports_list_view(request):
#     report_type = request.GET.get('type', '').upper().replace(' ', '_')
#     view_mode = request.GET.get('view', 'all')  # 'all', 'payments', 'receipts'

#     LOAN_TYPES = ['MTL LOAN', 'FDL LOAN', 'KVP/NSC LOAN']
#     DEPOSIT_TYPES = ['FIXED DEPOSITS', 'THRIFT FUNDS', 'WELFARE COLLECTIONS']
#     OTHER_TYPES = ['ADMISSION FEES', 'OTHER RECEIPTS', 'CASH WITHDRAWALS']

#     ALL_TYPES = LOAN_TYPES + DEPOSIT_TYPES + OTHER_TYPES
#     type_map = {t.upper().replace(' ', '_'): t for t in ALL_TYPES}

#     if report_type not in type_map:
#         return render(request, 'reports_list.html', {'error': f"Unknown report type: {report_type}"})

#     selected_type = type_map[report_type]
#     rows = []

#     if view_mode == 'payments':
#         # Use same logic as download_payments
#         payments = list(Loan.objects.all().order_by('-created_at', '-id'))
#         transactions = list(LoanRepayment.objects.select_related('loan').order_by('-created_at', '-id'))

#         new_payments = []
#         for t in transactions:
#             lt = get_loan_type(t)
#             if lt in SWAP_LOAN_TYPES:
#                 new_payments.append(t)
#             else:
#                 new_payments.append(t)  # keep all transactions as payments?

#         for p in payments:
#             lt = get_loan_type(p)
#             if lt not in SWAP_LOAN_TYPES:
#                 new_payments.append(p)

#         payments = [p for p in new_payments if get_loan_type(p).upper() == report_type]

#         for p in payments:
#             if hasattr(p, 'loan'):  # LoanRepayment
#                 loan_obj = p.loan
#                 rows.append({
#                     'gen_no': getattr(loan_obj, 'gen_no', '-'),
#                     'name': getattr(loan_obj, 'name', '-'),
#                     'ref': getattr(loan_obj, 'code', '-'),
#                     'balance': None,
#                     'interest': None,
#                     'amount': safe_decimal(getattr(p, 'total_payment', 0)),
#                     'row_type': 'payment'
#                 })
#             else:  # Loan
#                 rows.append({
#                     'gen_no': getattr(p, 'gen_no', '-'),
#                     'name': getattr(p, 'name', '-'),
#                     'ref': getattr(p, 'code', '-'),
#                     'balance': safe_decimal(getattr(p, 'amount', 0)),
#                     'interest': safe_decimal(getattr(p, 'interest', 0)),
#                     'amount': None,
#                     'row_type': 'payment'
#                 })

#     elif view_mode == 'receipts':
#         # Use same logic as download_receipts
#         transactions = list(LoanRepayment.objects.select_related('loan').order_by('-created_at', '-id'))
#         payments = list(Loan.objects.all().order_by('-created_at', '-id'))

#         new_receipts = []
#         for t in transactions:
#             lt = get_loan_type(t)
#             if lt not in SWAP_LOAN_TYPES:
#                 new_receipts.append(t)
#         for p in payments:
#             lt = get_loan_type(p)
#             if lt in SWAP_LOAN_TYPES:
#                 new_receipts.append(p)

#         receipts = [r for r in new_receipts if get_loan_type(r).upper() == report_type]

#         for r in receipts:
#             if hasattr(r, 'loan'):  # LoanRepayment
#                 loan_obj = r.loan
#                 rows.append({
#                     'gen_no': getattr(loan_obj, 'gen_no', '-'),
#                     'name': getattr(loan_obj, 'name', '-'),
#                     'ref': getattr(loan_obj, 'code', '-'),
#                     'balance': None,
#                     'interest': None,
#                     'amount': safe_decimal(getattr(r, 'total_payment', 0)),
#                     'row_type': 'receipt'
#                 })
#             else:  # Loan
#                 rows.append({
#                     'gen_no': getattr(r, 'gen_no', '-'),
#                     'name': getattr(r, 'name', '-'),
#                     'ref': getattr(r, 'code', '-'),
#                     'balance': safe_decimal(getattr(r, 'amount', 0)),
#                     'interest': safe_decimal(getattr(r, 'interest', 0)),
#                     'amount': None,
#                     'row_type': 'receipt'
#                 })

#     else:
#         # Default MTL loan view
#         loans_qs = Loan.objects.filter(type_of_loan=selected_type).order_by('gen_no', 'id')
#         for loan in loans_qs:
#             agg = LoanRepayment.objects.filter(loan=loan).aggregate(
#                 total_paid_principal=Coalesce(Sum('paid_to_principal'), Value(0), output_field=DecimalField()),
#                 total_paid_interest=Coalesce(Sum('paid_to_interest'), Value(0), output_field=DecimalField())
#             )
#             paid_principal = safe_decimal(agg['total_paid_principal'])
#             paid_interest = safe_decimal(agg.get('total_paid_interest', 0))
#             amount = safe_decimal(loan.amount)
#             interest = safe_decimal(loan.interest or 0)
#             balance = amount - paid_principal

#             if balance <= 0 and (interest - paid_interest <= 0 if selected_type in LOAN_TYPES + DEPOSIT_TYPES else True):
#                 continue

#             rows.append({
#                 'gen_no': loan.gen_no,
#                 'name': loan.name,
#                 'ref': loan.code or '',
#                 'balance': balance if selected_type in LOAN_TYPES + DEPOSIT_TYPES else None,
#                 'interest': (interest - paid_interest) if selected_type in LOAN_TYPES + DEPOSIT_TYPES else None,
#                 'amount': balance if selected_type in OTHER_TYPES else None,
#                 'row_type': 'default'
#             })

#     context = {
#         'report_type': selected_type,
#         'rows': rows,
#         'is_loan_or_deposit': selected_type in LOAN_TYPES + DEPOSIT_TYPES,
#         'is_other': selected_type in OTHER_TYPES,
#         'view_mode': view_mode
#     }

#     return render(request, 'reports_list.html', context)
# 2002-dec

from django.shortcuts import render
from django.http import HttpResponse
import xlsxwriter
from .models import Loan, LoanRepayment
from decimal import Decimal

SWAP_LOAN_TYPES = ['FIXED DEPOSITS', 'THRIFT FUNDS', 'WELFARE COLLECTIONS']

def safe_decimal(value):
    try:
        return Decimal(value or 0)
    except:
        return Decimal(0)

def download_reports(request):
    # Fetch all transactions (LoanRepayment)
    transactions = list(LoanRepayment.objects.select_related('loan').order_by('-created_at', '-id'))
    # Fetch all payments (Loan)
    payments = list(Loan.objects.all().order_by('-created_at', '-id'))

    # ---------------- Swap Logic ----------------
    new_transactions, new_payments = [], []
    for t in transactions:
        loan_type = t.loan.type_of_loan
        if loan_type in SWAP_LOAN_TYPES:
            new_payments.append(t)
        else:
            new_transactions.append(t)

    for p in payments:
        if p.type_of_loan in SWAP_LOAN_TYPES:
            new_transactions.append(p)
        else:
            new_payments.append(p)

    transactions = new_transactions
    payments = new_payments

    return render(request, 'download_reports.html', {
        'transactions': transactions,
        'payments': payments,
    })

def get_loan_type(obj):
    """Return loan type from Loan or LoanRepayment safely."""
    if hasattr(obj, "loan"):          # LoanRepayment
        return obj.loan.type_of_loan
    if hasattr(obj, "type_of_loan"):  # Loan
        return obj.type_of_loan
    return "-"

def get_amount(obj):
    """Return amount safely."""
    if hasattr(obj, "amount"):
        return obj.amount
    if hasattr(obj, "total_payment"):
        return obj.total_payment
    return 0
    
from django.http import HttpResponse
import xlsxwriter
from .models import Loan, LoanRepayment

SWAP_LOAN_TYPES = ['FIXED DEPOSITS', 'THRIFT FUNDS', 'WELFARE COLLECTIONS']

def safe_decimal(value):
    """Convert value to Decimal safely, treating invalid strings as 0."""
    from decimal import Decimal, InvalidOperation
    try:
        return Decimal(value)
    except (InvalidOperation, TypeError, ValueError):
        return Decimal('0')

def get_loan_type(obj):
    """Safely return loan type from Loan or LoanRepayment."""
    if hasattr(obj, "loan"):  # LoanRepayment
        return getattr(obj.loan, "type_of_loan", "-")
    return getattr(obj, "type_of_loan", "-")  # Loan

def download_payments(request, loan_type=None):
    loan_type_filter = request.GET.get("type", loan_type)

    payments = list(Loan.objects.all().order_by('-created_at', '-id'))
    transactions = list(LoanRepayment.objects.select_related('loan').order_by('-created_at', '-id'))

    new_transactions, new_payments = [], []

    # Swap logic
    for t in transactions:
        lt = get_loan_type(t)
        if lt in SWAP_LOAN_TYPES:
            new_payments.append(t)
        else:
            new_transactions.append(t)

    for p in payments:
        lt = get_loan_type(p)
        if lt in SWAP_LOAN_TYPES:
            new_transactions.append(p)
        else:
            new_payments.append(p)

    payments = new_payments

    # Apply loan_type filter
    if loan_type_filter:
        loan_type_filter = loan_type_filter.upper()
        payments = [p for p in payments if get_loan_type(p).upper() == loan_type_filter]

    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = f'attachment; filename="payments_{loan_type_filter or "all"}.xlsx"'

    workbook = xlsxwriter.Workbook(response, {'in_memory': True})
    ws = workbook.add_worksheet("Payments")

    headers = ['Date', 'Code', 'Loan Type', 'Cash', 'Bank1', 'Bank2', 'Adjustment', 'Amount']
    for col, h in enumerate(headers):
        ws.write(0, col, h)

    for row_num, p in enumerate(payments, 1):
        ws.write(row_num, 0, p.created_at.strftime('%d-%m-%Y'))
        ws.write(row_num, 1, getattr(p, 'code', getattr(p, 'loan_id', '-')))
        ws.write(row_num, 2, get_loan_type(p))
        ws.write(row_num, 3, safe_decimal(getattr(p, 'cash', 0)))
        ws.write(row_num, 4, safe_decimal(getattr(p, 'bank1', 0)))
        ws.write(row_num, 5, safe_decimal(getattr(p, 'bank2', 0)))
        ws.write(row_num, 6, safe_decimal(getattr(p, 'adj', 0)))
        # Use total_payment if LoanRepayment doesn't have 'amount'
        ws.write(row_num, 7, safe_decimal(getattr(p, 'amount', getattr(p, 'total_payment', 0))))

    workbook.close()
    return response


def download_receipts(request, loan_type=None):
    loan_type_filter = request.GET.get("type", loan_type)

    transactions = list(LoanRepayment.objects.select_related('loan').order_by('-created_at', '-id'))
    payments = list(Loan.objects.all().order_by('-created_at', '-id'))

    new_transactions, new_payments = [], []

    # Swap logic
    for t in transactions:
        lt = get_loan_type(t)
        if lt in SWAP_LOAN_TYPES:
            new_payments.append(t)
        else:
            new_transactions.append(t)

    for p in payments:
        lt = get_loan_type(p)
        if lt in SWAP_LOAN_TYPES:
            new_transactions.append(p)
        else:
            new_payments.append(p)

    transactions = new_transactions

    # Apply loan_type filter
    if loan_type_filter:
        loan_type_filter = loan_type_filter.upper()
        transactions = [t for t in transactions if get_loan_type(t).upper() == loan_type_filter]

    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = f'attachment; filename="receipts_{loan_type_filter or "all"}.xlsx"'

    workbook = xlsxwriter.Workbook(response, {'in_memory': True})
    ws = workbook.add_worksheet("Receipts")

    headers = ['Date', 'Loan Id', 'Loan Type', 'Cash', 'Bank1', 'Bank2', 'Adjustment']
    for col, h in enumerate(headers):
        ws.write(0, col, h)

    for row_num, t in enumerate(transactions, 1):
        ws.write(row_num, 0, t.created_at.strftime('%d-%m-%Y'))
        ws.write(row_num, 1, getattr(t, 'loan_id', '-'))
        ws.write(row_num, 2, get_loan_type(t))
        ws.write(row_num, 3, safe_decimal(getattr(t, 'cash', 0)))
        ws.write(row_num, 4, safe_decimal(getattr(t, 'bank1', 0)))
        ws.write(row_num, 5, safe_decimal(getattr(t, 'bank2', 0)))
        ws.write(row_num, 6, safe_decimal(getattr(t, 'adj', 0)))

    workbook.close()
    return response


from decimal import Decimal
from django.shortcuts import render
from django.db.models import Sum, Value, DecimalField
from django.db.models.functions import Coalesce
from .models import Loan, LoanRepayment

SWAP_LOAN_TYPES = ['FIXED DEPOSITS', 'THRIFT FUNDS', 'WELFARE COLLECTIONS']


def fmt(dt):
    if dt:
        return dt.strftime("%Y-%m-%d %H:%M:%S")
    return "-"


# ---------- SAFE DECIMAL ----------
def safe_decimal(value):
    try:
        return Decimal(value or 0)
    except:
        return Decimal('0')


# ---------- NORMALIZED TYPE ----------
def normalize_type(t):
    return t.upper().replace(" ", "_").replace("/", "_").replace("-", "_")


def get_loan_type(obj):
    if hasattr(obj, "loan") and obj.loan:
        t = obj.loan.type_of_loan
    else:
        t = getattr(obj, "type_of_loan", "-")

    return normalize_type(t)





# ---------- FIXED CASH BOOK SWAP LOGIC ----------
def cash_book_swap_logic():
    """
    Correct classification:
    - LoanRepayment = ALWAYS RECEIPT
    - Loan = ALWAYS PAYMENT
    - Swap only the 3 deposit types
    """

    repayments = list(LoanRepayment.objects.select_related('loan').order_by('-created_at', '-id'))
    loans = list(Loan.objects.all().order_by('-created_at', '-id'))

    payments = []   # Loan objects
    receipts = []   # LoanRepayment objects

    for r in repayments:
        loan_type = r.loan.type_of_loan
        if loan_type in SWAP_LOAN_TYPES:
            payments.append(r)
        else:
            receipts.append(r)

    for l in loans:
        loan_type = l.type_of_loan
        if loan_type in SWAP_LOAN_TYPES:
            receipts.append(l)
        else:
            payments.append(l)

    return payments, receipts


# ---------- MAIN VIEW ----------
def reports_list_view(request):
    report_type = normalize_type(request.GET.get('type', ''))
    view_mode = request.GET.get('view', 'all')

    LOAN_TYPES = ['MTL LOAN', 'FDL LOAN', 'KVP/NSC LOAN']
    DEPOSIT_TYPES = ['FIXED DEPOSITS', 'THRIFT FUNDS', 'WELFARE COLLECTIONS']
    OTHER_TYPES = ['ADMISSION FEES', 'OTHER RECEIPTS', 'CASH WITHDRAWALS']

    ALL_TYPES = LOAN_TYPES + DEPOSIT_TYPES + OTHER_TYPES
    type_map = {normalize_type(t): t for t in ALL_TYPES}

    if report_type not in type_map:
        return render(request, 'reports_list.html', {'error': f"Unknown report type: {report_type}"})

    selected_type = type_map[report_type]
    rows = []

    payments, receipts = cash_book_swap_logic()

    # ---------------------------------------------------
    # PAYMENTS VIEW
    # ---------------------------------------------------
    if view_mode == 'payments':

        filtered = [p for p in payments if get_loan_type(p) == report_type]

        for p in filtered:
            adj = safe_decimal(getattr(p, "adjustment", 0))

            if hasattr(p, "loan") and p.loan:
                ln = p.loan
                rows.append({
                    'gen_no': ln.gen_no,
                    'name': ln.name,
                    'ref': ln.code,
                    'cash': safe_decimal(p.cash),
                    'bank1': safe_decimal(p.bank1),
                    'bank2': safe_decimal(p.bank2),
                    'adjustment': adj,
                    'created_at': p.created_at.strftime("%Y-%m-%d %H:%M:%S") if hasattr(p, "created_at") else "-"
                })
            else:
                rows.append({
                    'gen_no': p.gen_no,
                    'name': p.name,
                    'ref': p.code,
                    'cash': safe_decimal(p.amount),
                    'bank1': Decimal('0'),
                    'bank2': Decimal('0'),
                    'adjustment': Decimal('0'),
                    'created_at': p.created_at.strftime("%Y-%m-%d %H:%M:%S") if hasattr(p, "created_at") else "-"
                })

    # ---------------------------------------------------
    # RECEIPTS VIEW
    # ---------------------------------------------------
    elif view_mode == 'receipts':

        filtered = [r for r in receipts if get_loan_type(r) == report_type]

        for r in filtered:
            adj = safe_decimal(getattr(r, "adjustment", 0))

            if hasattr(r, "loan") and r.loan:
                ln = r.loan
                rows.append({
                    'gen_no': ln.gen_no,
                    'name': ln.name,
                    'ref': ln.code,
                    'cash': safe_decimal(r.cash),
                    'bank1': safe_decimal(r.bank1),
                    'bank2': safe_decimal(r.bank2),
                    'adjustment': adj,
                    'created_at': r.created_at.strftime("%Y-%m-%d %H:%M:%S")
                    
                })
            else:
                rows.append({
                    'gen_no': r.gen_no,
                    'name': r.name,
                    'ref': r.code,
                    'cash': safe_decimal(r.amount),
                    'bank1': Decimal('0'),
                    'bank2': Decimal('0'),
                    'adjustment': Decimal('0'),
                    'created_at': r.created_at.strftime("%Y-%m-%d %H:%M:%S")
                })

    # ---------------------------------------------------
    # ALL VIEW (DEFAULT)
    # ---------------------------------------------------
    else:
        loans_qs = Loan.objects.filter(type_of_loan=selected_type).order_by('gen_no', 'id')

        for loan in loans_qs:
            if selected_type in OTHER_TYPES:
                rows.append({
                    'gen_no': loan.gen_no,
                    'name': loan.name,
                    'ref': loan.code,
                    'balance': safe_decimal(loan.amount or 0),
                    'interest': None,
                    "created_at": fmt(getattr(loan, 'created_at', None)),
                })
                continue

            agg = LoanRepayment.objects.filter(loan=loan).aggregate(
                total_paid_principal=Coalesce(Sum('paid_to_principal'), Value(0), output_field=DecimalField()),
                total_paid_interest=Coalesce(Sum('paid_to_interest'), Value(0), output_field=DecimalField())
            )

            paid_p = safe_decimal(agg['total_paid_principal'])
            paid_i = safe_decimal(agg['total_paid_interest'])

            balance = safe_decimal(loan.amount) - paid_p
            interest_bal = safe_decimal(loan.interest) - paid_i

            rows.append({
                'gen_no': loan.gen_no,
                'name': loan.name,
                'ref': loan.code,
                'balance': balance,
                'interest': interest_bal,
                "created_at": fmt(getattr(loan, 'created_at', None)),
            })

    return render(request, 'reports_list.html', {
        'report_type': selected_type,
        'rows': rows,
        'is_loan_or_deposit': selected_type not in OTHER_TYPES,
        'view_mode': view_mode
    })



# fun
from django.http import HttpResponse
import csv
from decimal import Decimal
from .models import Loan, LoanRepayment

def safe_decimal(value):
    try:
        return Decimal(value or 0)
    except:
        return Decimal('0')

def download_report_view(request, report_type, view_mode):
    search_term = request.GET.get('search', '').lower()
    rows = []

    if view_mode == 'all':
        loans = Loan.objects.all()
        for loan in loans:
            ref_value = getattr(loan, 'type_of_loan', '')
            row = {
                'gen_no': getattr(loan, 'gen_no', ''),
                'name': getattr(loan, 'name', ''),
                'ref': ref_value,
                'balance': safe_decimal(getattr(loan, 'balance', 0)),
                'interest': safe_decimal(getattr(loan, 'interest', 0)),
            }
            if search_term in row['name'].lower() or search_term in str(row['ref']).lower():
                rows.append(row)

    elif view_mode == 'receipts':
        repayments = LoanRepayment.objects.filter(total_payment__gt=0)
        for r in repayments:
            row = {
                'gen_no': getattr(r.loan, 'gen_no', ''),
                'name': getattr(r.loan, 'name', ''),
                'ref': getattr(r.loan, 'type_of_loan', ''),
                'cash': safe_decimal(r.cash),
                'bank1': safe_decimal(r.bank1),
                'bank2': safe_decimal(r.bank2),
                'adjustment': safe_decimal(r.adj),
            }
            if search_term in row['name'].lower() or search_term in str(row['ref']).lower():
                rows.append(row)

    else:  # payments
        repayments = LoanRepayment.objects.filter(paid_to_principal__gt=0)
        for r in repayments:
            row = {
                'gen_no': getattr(r.loan, 'gen_no', ''),
                'name': getattr(r.loan, 'name', ''),
                'ref': getattr(r.loan, 'type_of_loan', ''),
                'cash': safe_decimal(r.cash),
                'bank1': safe_decimal(r.bank1),
                'bank2': safe_decimal(r.bank2),
                'adjustment': safe_decimal(r.adj),
            }
            if search_term in row['name'].lower() or search_term in str(row['ref']).lower():
                rows.append(row)

    # Generate CSV
    response = HttpResponse(content_type='text/csv')
    filename = f"{report_type}_{view_mode}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)

    if view_mode == 'all':
        writer.writerow(['Gen No', 'Name', 'Ref', 'Balance', 'Interest'])
        for row in rows:
            writer.writerow([row['gen_no'], row['name'], row['ref'], row['balance'], row['interest']])
    else:
        writer.writerow(['Gen No', 'Name', 'Ref', 'Cash', 'Bank1', 'Bank2', 'Adjustment'])
        for row in rows:
            writer.writerow([row['gen_no'], row['name'], row['ref'], row['cash'], row['bank1'], row['bank2'], row['adjustment']])

    return response



# day last





# going home
# views.py
import json
from datetime import datetime
from django.http import HttpResponse
from openpyxl import Workbook

def download_receipts_dynamic(request):
    if request.method != 'POST':
        return HttpResponse(status=400)
    data = json.loads(request.body or '{}')
    rows = data.get('rows', [])

    wb = Workbook()
    ws = wb.active
    ws.title = "Receipts"
    ws.append(["Date", "Loan Id", "Loan Type", "Cash", "Bank 1", "Bank 2", "Adjustment"])

    for row in rows:
        ws.append(row)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=receipts_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(response)
    return response


def download_payments_dynamic(request):
    if request.method != 'POST':
        return HttpResponse(status=400)
    data = json.loads(request.body or '{}')
    rows = data.get('rows', [])

    wb = Workbook()
    ws = wb.active
    ws.title = "Payments"
    ws.append(["Date", "Loan Id", "Loan Type", "Cash", "Bank 1", "Bank 2", "Adjustment"])

    for row in rows:
        ws.append(row)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=payments_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(response)
    return response












from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt

@csrf_exempt
def user_search(request):
    q = request.GET.get('q', '').strip()

    users = User.objects.filter(code__icontains=q).values('code','name')[:10]

    return JsonResponse(list(users), safe=False)








# last day
# # from openpyxl import load_workbook
# # from decimal import Decimal
# # from django.shortcuts import render
# # from .models import User, Loan, CashEntry
# # from uuid import uuid4

# # # ----------------- UTILITY FUNCTIONS -----------------
# # def clean(v):
# #     return str(v).strip() if v not in [None, "", " "] else ""

# # def num(v):
# #     try:
# #         return Decimal(v) if v not in [None, "", " "] else Decimal("0.00")
# #     except:
# #         return Decimal("0.00")

# # # ----------------- UPLOAD EXCEL -----------------
# # def upload_excel(request):
# #     inserted = {"Users":0,"CashEntry":0,"Loans":0,"Deposits":0,"Others":0}
# #     updated  = {"Users":0,"CashEntry":0,"Loans":0,"Deposits":0,"Others":0}
# #     errors = []

# #     if request.method == "POST":
# #         excel_file = request.FILES.get("excel_file")
# #         if not excel_file:
# #             return render(request,"upload_excel.html",{"error":"No Excel selected!"})

# #         try:
# #             wb = load_workbook(excel_file, data_only=True)
# #         except Exception as e:
# #             return render(request,"upload_excel.html",{"error":f"Invalid Excel: {e}"})

# #         # ---------------- 1) USERS ----------------
# #         if "User" in wb.sheetnames:
# #             ws = wb["User"]
# #             headers = [c.value for c in ws[1]]
# #             for r, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
# #                 data = dict(zip(headers,row))
# #                 name = clean(data.get("name"))
# #                 if not name:
# #                     continue

# #                 payload = {
# #                     "name": name,
# #                     "Mobile": clean(data.get("Mobile")),
# #                     "Address": clean(data.get("Address")),
# #                     "IFSCcode": clean(data.get("IFSCcode")),
# #                     "AccountNo1": num(data.get("AccountNo1")),
# #                     "AccountNo2": num(data.get("AccountNo2")),
# #                     "IFSCcode2": clean(data.get("IFSCcode2")),
# #                     "Age": clean(data.get("Age")),
# #                     "Email": clean(data.get("Email")),
# #                     "BankAccountName": clean(data.get("BankAccountName")),
# #                     "BranchName": clean(data.get("BranchName")),
# #                     "BankAccountName2": clean(data.get("BankAccountName2")),
# #                     "BranchName2": clean(data.get("BranchName2")),
# #                 }

# #                 try:
# #                     obj = None
# #                     if data.get("code"):
# #                         obj = User.objects.filter(code=clean(data["code"])).first()
# #                     if not obj:
# #                         obj = User.objects.filter(name__iexact=name).first()

# #                     if obj:
# #                         for k,v in payload.items(): setattr(obj,k,v)
# #                         obj.save()
# #                         updated["Users"] += 1
# #                     else:
# #                         User.objects.create(**payload)
# #                         inserted["Users"] += 1
# #                 except Exception as e:
# #                     errors.append(f"User row {r} error: {e}")

# #         # ---------------- 2) CASH ENTRY ----------------
# #         if "AddCash" in wb.sheetnames:
# #             ws = wb["AddCash"]
# #             headers = [c.value for c in ws[1]]
# #             for r,row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
# #                 data = dict(zip(headers,row))
# #                 type_of_cash = clean(data.get("type_of_cash"))
# #                 if type_of_cash not in ["Cash", "Bank 1", "Bank 2", "Adjustment"]:
# #                     errors.append(f"AddCash row {r}: Invalid type_of_cash '{type_of_cash}'")
# #                     continue

# #                 payload = {
# #                     "amount": num(data.get("amount")),
# #                     "type_of_cash": type_of_cash,
# #                     "remarks": clean(data.get("remarks")),
# #                     "datetime": data.get("datetime") if data.get("datetime") else None,
# #                     "code": str(uuid4())[:8]  # ensure unique code
# #                 }
# #                 try:
# #                     exists = CashEntry.objects.filter(
# #                         amount=payload["amount"],
# #                         type_of_cash=payload["type_of_cash"],
# #                         remarks=payload["remarks"]
# #                     ).first()
# #                     if exists:
# #                         updated["CashEntry"] += 1
# #                     else:
# #                         CashEntry.objects.create(**payload)
# #                         inserted["CashEntry"] += 1
# #                 except Exception as e:
# #                     errors.append(f"CashEntry row {r} error: {e}")

# #         # ---------------- 3) LOANS / DEPOSITS / OTHERS ----------------
# #         SHEETS = {
# #             "Loans": ["MTL LOAN","FDL LOAN","KVP/NSC LOAN"],
# #             "Deposits": ["FIXED DEPOSITS","THRIFT FUNDS","WELFARE COLLECTIONS"],
# #             "Others": ["ADMISSION FEES","OTHER RECEIPTS","CASH WITHDRAWALS"]
# #         }

# #         for sheet, types in SHEETS.items():
# #             if sheet in wb.sheetnames:
# #                 ws = wb[sheet]
# #                 headers = [c.value for c in ws[1]]
# #                 for r,row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
# #                     data = dict(zip(headers,row))
# #                     type_of_loan = clean(data.get("type_of_loan"))
# #                     if type_of_loan not in types:
# #                         continue

# #                     name = clean(data.get("name"))
# #                     amount = num(data.get("amount"))
# #                     balance = num(data.get("balance"))
# #                     interest = num(data.get("interest")) if sheet in ["Loans", "Deposits"] else Decimal("0.00")
# #                     type_of_cash = clean(data.get("type_of_cash"))
# #                     if type_of_cash not in ["Cash", "Bank 1", "Bank 2", "Adjustment"]:
# #                         errors.append(f"{sheet} row {r}: Invalid type_of_cash '{type_of_cash}'")
# #                         continue

# #                     # Match user by name
# #                     user_obj = User.objects.filter(name__iexact=name).first()
# #                     if not user_obj:
# #                         errors.append(f"{sheet} row {r}: user '{name}' not found")
# #                         continue

# #                     payload = {
# #                         "gen_no": user_obj.code,
# #                         "name": name,
# #                         "amount": amount,
# #                         "balance": balance,
# #                         "interest": interest,
# #                         "type_of_loan": type_of_loan,
# #                         "type_of_cash": type_of_cash
# #                     }

# #                     try:
# #                         obj = Loan.objects.filter(gen_no=user_obj.code, type_of_loan=type_of_loan).first()
# #                         if obj:
# #                             for k,v in payload.items(): setattr(obj,k,v)
# #                             obj.save()
# #                             updated[sheet] += 1
# #                         else:
# #                             Loan.objects.create(**payload)
# #                             inserted[sheet] += 1
# #                     except Exception as e:
# #                         errors.append(f"{sheet} row {r} error: {e}")

# #         # ---------------- FINISH ----------------
# #         return render(request,"upload_excel.html",{
# #             "inserted": inserted,
# #             "updated": updated,
# #             "errors": errors
# #         })

#     # GET request
#     return render(request,"upload_excel.html")
from decimal import Decimal
from django.shortcuts import render
from django.utils import timezone
from openpyxl import load_workbook
from uuid import uuid4
from .models import User, Loan, CashEntry

# ---------------- UTILITY FUNCTIONS ----------------
def clean(v): 
    return str(v).strip() if v not in [None, "", " "] else ""

def num(v):
    try: 
        return Decimal(v) if v not in [None, "", " "] else Decimal("0.00")
    except: 
        return Decimal("0.00")

# # ---------------- UPLOAD EXCEL ----------------
# def upload_excel(request):
#     inserted = {"Users":0,"CashEntry":0,"Loans":0,"Deposits":0,"Others":0}
#     updated  = {"Users":0,"CashEntry":0,"Loans":0,"Deposits":0,"Others":0}
#     errors = []

#     if request.method == "POST":
#         file = request.FILES.get("excel_file")
#         if not file:
#             return render(request, "upload_excel.html", {"error": "No File Selected"})

#         try:
#             wb = load_workbook(file, data_only=True)
#         except Exception as e:
#             return render(request, "upload_excel.html", {"error": f"Invalid File: {e}"})

#         # ---------------- USERS ----------------
#         # ---------------- USERS ----------------
#         if "Users" in wb.sheetnames:
#             ws = wb["Users"]
#             headers = [str(c.value).strip().lower() for c in ws[1]]

#             for r, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
#                 data = dict(zip(headers, row))

#                 gen_no = clean(data.get("gen no"))
#                 name = clean(data.get("name"))
#                 mobile = clean(data.get("mobile"))
#                 address = clean(data.get("address"))

#                 if not name:
#                     continue  # name is mandatory

#                 try:
#                     # CASE 1: Gen No provided
#                     if gen_no:
#                         user = User.objects.filter(code=gen_no).first()
#                         if user:
#                             user.name = name
#                             user.Mobile = mobile
#                             user.Address = address
#                             user.save()
#                             updated["Users"] += 1
#                         else:
#                             User.objects.create(
#                                 code=gen_no,
#                                 name=name,
#                                 Mobile=mobile,
#                                 Address=address
#                             )
#                             inserted["Users"] += 1

#                     # CASE 2: Gen No NOT provided
#                     else:
#                         user = User.objects.filter(name__iexact=name).first()
#                         if user:
#                             user.Mobile = mobile
#                             user.Address = address
#                             user.save()
#                             updated["Users"] += 1
#                         else:
#                             User.objects.create(
#                                 name=name,
#                                 Mobile=mobile,
#                                 Address=address
#                             )
#                             # CNxxxx auto-generated
#                             inserted["Users"] += 1

#                 except Exception as e:
#                     errors.append(f"Users row {r} error: {e}")


#  # ---------------- LOANS ONLY ----------------
#                 if "Loans" in wb.sheetnames:
#                     ws = wb["Loans"]

#                     # CLEAN HEADERS: strip, lowercase, replace spaces with _
#                     headers = []
#                     for c in ws[1]:
#                         if c.value:
#                             headers.append(str(c.value).strip().lower().replace(" ", "_"))
#                         else:
#                             headers.append("unknown")  # fallback for empty header

#                     for r, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
#                         data = dict(zip(headers, row))

#                         name = clean(data.get("name"))
#                         type_of_loan = clean(data.get("type_of_loan"))
#                         excel_code = clean(data.get("code"))

#                         if not name or not type_of_loan:
#                             print(f"Skipping row {r} → Missing name/type_of_loan")
#                             continue

#                         type_of_loan = type_of_loan.strip().upper()
#                         if type_of_loan not in ["MTL LOAN", "FDL LOAN", "KVP/NSC LOAN"]:
#                             print(f"Skipping row {r} → Invalid type_of_loan: {type_of_loan}")
#                             continue

#                         # Convert amounts safely
#                         cash = num(data.get("cash"))
#                         bank1 = num(data.get("bank1"))
#                         bank2 = num(data.get("bank2"))
#                         adj = num(data.get("adjustment"))
#                         interest = num(data.get("interest"))
#                         balance = num(data.get("balance"))
#                         total_paid = cash + bank1 + bank2 + adj

#                         # Find or create user
#                         user = User.objects.filter(name__iexact=name).first()
#                         if not user:
#                             user = User.objects.create(name=name, Mobile="", Address="")

#                         payload = {
#                             "gen_no": user.code,
#                             "name": name,
#                             "type_of_loan": type_of_loan,
#                             "amount": total_paid,
#                             "interest": interest,
#                             "balance": balance,
#                             "cash": str(cash),
#                             "bank1": str(bank1),
#                             "bank2": str(bank2),
#                             "adj": str(adj),
#                         }
#                         if excel_code:
#                             payload["code"] = excel_code

#                         try:
#                             obj = Loan.objects.filter(gen_no=user.code, type_of_loan=type_of_loan).first()
#                             if obj:
#                                 for k, v in payload.items():
#                                     setattr(obj, k, v)
#                                 obj.save()
#                                 updated["Loans"] += 1
#                                 print(f"Updated Loan row {r}")
#                             else:
#                                 Loan.objects.create(**payload)
#                                 inserted["Loans"] += 1
#                                 print(f"Inserted Loan row {r}")
#                         except Exception as e:
#                             errors.append(f"Loans row {r} error: {e}")
#                             print(f"Error row {r}: {e}")

#         return render(request, "upload_excel.html", {
#             "inserted": inserted,
#             "updated": updated,
#             "errors": errors
#         })

#     # ---------------- GET ----------------
#     return render(request, "upload_excel.html")



# from django.http import HttpResponse
# from openpyxl import Workbook

# def download_sample_excel(request):
#     wb = Workbook()
    
#     # ---------------- USERS ----------------
#     ws = wb.active
#     ws.title = "User"
#     user_headers = [
#         "name", "Mobile", "Address", "IFSCcode", "AccountNo1", "AccountNo2", "IFSCcode2",
#         "Age", "Email", "BankAccountName", "BranchName", "BankAccountName2", "BranchName2"
#     ]
#     ws.append(user_headers)
    
#     # ---------------- LOANS ----------------
#     def add_sheet(title, include_interest=True):
#         ws = wb.create_sheet(title)
#         headers = ["name", "type_of_loan", "Cash", "Bank1", "Bank2", "Adjustment", "balance"]
#         if include_interest:
#             headers.insert(-1, "interest")  # interest before balance
#         ws.append(headers)
    
#     add_sheet("Loans", include_interest=True)
#     add_sheet("Deposits", include_interest=True)
#     add_sheet("Others", include_interest=False)
    
#     # ---------------- CASH ENTRY ----------------
#     ws = wb.create_sheet("AddCash")
#     cash_headers = ["amount", "type_of_cash", "remarks", "datetime (optional)"]
#     ws.append(cash_headers)
    
#     # ---------------- RESPONSE ----------------
#     response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
#     response['Content-Disposition'] = 'attachment; filename=empty_sample_upload.xlsx'
#     wb.save(response)
#     return response


















from django.http import HttpResponse
from openpyxl import Workbook


def download_sample_excel(request):
    """
    Generate a sample Excel file with sheets:
    Users, Loans, Deposits, Others
    aligned exactly with upload logic.
    """

    wb = Workbook()

    # ================= USERS =================
    ws_users = wb.active
    ws_users.title = "Users"
    ws_users.append(["Gen No", "Name", "Mobile", "Address"])



    # ================= LOANS =================
    ws_loans = wb.create_sheet("Loans")
    ws_loans.append([
        "Name",
        "Type of Loan",
        "Cash",
        "Bank1",
        "Bank2",
        "Adjustment",
        "Interest",
        "Balance",
        "Code"
    ])



    # ================= DEPOSITS =================
    ws_deposits = wb.create_sheet("Deposits")
    ws_deposits.append([
        "Name",
        "Type of Loan",
        "Cash",
        "Bank1",
        "Bank2",
        "Adjustment",
        "Interest",
        "Balance",
        "Code"
    ])



    # ================= OTHERS =================
    ws_others = wb.create_sheet("Others")
    ws_others.append([
        "Name",
        "Transaction Type",
        "Type of Loan",
        "Cash",
        "Bank1",
        "Bank2",
        "Code"
    ])



    # ================= RESPONSE =================
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=LoanHub_Sample.xlsx"

    wb.save(response)
    return response



from django.http import HttpResponse
import pandas as pd
from .models import LoanRepayment, CashEntry   # change if different

def download_reports_payments(request):
    payments = LoanRepayment.objects.all().values()   # filter later if needed
    df = pd.DataFrame(payments)

    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="Payments_Report.xlsx"'
    df.to_excel(response, index=False)
    return response


def download_reports_receipts(request):
    receipts = CashEntry.objects.all().values()   # filter later if needed
    df = pd.DataFrame(receipts)

    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="Receipts_Report.xlsx"'
    df.to_excel(response, index=False)
    return response






from django.contrib.auth import logout
from django.shortcuts import redirect

def logout_view(request):
    logout(request)
    return redirect('login')  # redirect to your login page






# views.py
# views.py
from django.http import JsonResponse
from .models import User

def search_user(request):
    q = request.GET.get('q', '').strip()
    
    if q.upper() == 'LH' or q == '':
        users = User.objects.all()  # return all users
    else:
        users = User.objects.filter(name__icontains=q) | User.objects.filter(code__icontains=q)

    user_list = [{'code': u.code, 'name': u.name} for u in users]
    return JsonResponse(user_list, safe=False)


# views.py
from django.http import JsonResponse
from .models import User

from django.http import JsonResponse
from django.db.models import Q
from .models import User

def search_user_codes(request):
    query = request.GET.get("q") or request.GET.get("term") or ""

    if query.lower() == "lh" or query.strip() == "":
        # Show all users when input clicked or query is 'lh'
        users = User.objects.all().order_by("code")[:200]  # limit to 200 for safety
    else:
        # Filter by code OR name
        users = User.objects.filter(
            Q(code__icontains=query) | Q(name__icontains=query)
        ).order_by("code")[:200]  # limit for performance

    data = [{"code": u.code, "name": u.name} for u in users]
    return JsonResponse(data, safe=False)





from django.http import JsonResponse
from .models import User

def all_users(request):
    users = User.objects.values('code', 'name')
    return JsonResponse(list(users), safe=False)




# date @5/01/2025
from decimal import Decimal
from django.shortcuts import render

def cash_withdrawals(request):

    # ------------------ Base data ------------------
    transactions = list(
        LoanRepayment.objects.select_related('loan').order_by('-created_at', '-id')
    )

    payments = list(
        Loan.objects.all().order_by('-created_at', '-id')
    )

    # ------------------ CashEntry (SAME AS cash_book) ------------------
    cash_entries = CashEntry.objects.all().order_by('-datetime')

    class CashBookEntry:
        def __init__(self, cash_entry):
            raw_amount = Decimal(cash_entry.amount)
            self.amount = abs(raw_amount)

            self.cash = self.amount if cash_entry.type_of_cash == "Cash" else Decimal('0')
            self.bank1 = self.amount if cash_entry.type_of_cash == "Bank1" else Decimal('0')
            self.bank2 = self.amount if cash_entry.type_of_cash == "Bank2" else Decimal('0')
            self.adj = self.amount if cash_entry.type_of_cash == "Adjustment" else Decimal('0')

            self.type_of_loan = cash_entry.type_of_loan
            self.loan = type('LoanObj', (object,), {
                'id': cash_entry.code,
                'type_of_loan': cash_entry.type_of_loan
            })()

            self.created_at = cash_entry.datetime
            self.is_negative = raw_amount < 0

    for entry in cash_entries:
        cb_entry = CashBookEntry(entry)

        if cb_entry.is_negative:
            payments.append(cb_entry)
        elif entry.type_of_cash in SWAP_LOAN_TYPES:
            payments.append(cb_entry)
        else:
            transactions.append(cb_entry)

    # ------------------ Swap Logic (SAME AS cash_book) ------------------
    new_transactions, new_payments = [], []

    for t in transactions:
        loan_type = getattr(t.loan, 'type_of_loan', None)
        if loan_type in SWAP_LOAN_TYPES:
            new_payments.append(t)
        else:
            new_transactions.append(t)

    for p in payments:
        if hasattr(p, 'loan'):
            loan_type = getattr(p.loan, 'type_of_loan', None)
        else:
            loan_type = getattr(p, 'type_of_loan', None)

        if loan_type in SWAP_LOAN_TYPES:
            new_transactions.append(p)
        else:
            new_payments.append(p)

    transactions = new_transactions
    payments = new_payments

    # ------------------ OtherCashTransaction (ADD – SAME AS cash_book) ------------------
    other_receipts = list(
        OtherCashTransaction.objects.filter(transaction_type='RECEIPT')
    )

    other_payments = list(
        OtherCashTransaction.objects.filter(transaction_type='PAYMENT')
    )

    transactions.extend(other_receipts)
    payments.extend(other_payments)

    # ------------------ Helpers ------------------
    def safe_decimal(value):
        try:
            return Decimal(value or 0)
        except:
            return Decimal(0)

    def get_adj(obj):
        return safe_decimal(obj.adj) if hasattr(obj, 'adj') else Decimal('0')

    # ------------------ Totals ------------------
    receipts_totals = {
        'cash': sum(safe_decimal(getattr(t, 'cash', 0)) for t in transactions),
        'bank1': sum(safe_decimal(getattr(t, 'bank1', 0)) for t in transactions),
        'bank2': sum(safe_decimal(getattr(t, 'bank2', 0)) for t in transactions),
        'adj': sum(get_adj(t) for t in transactions),
    }

    payments_totals = {
        'cash': sum(safe_decimal(getattr(p, 'cash', 0)) for p in payments),
        'bank1': sum(safe_decimal(getattr(p, 'bank1', 0)) for p in payments),
        'bank2': sum(safe_decimal(getattr(p, 'bank2', 0)) for p in payments),
        'adj': sum(get_adj(p) for p in payments),
    }

    # ------------------ Withdrawals ------------------
    withdrawals = {
        'cash': receipts_totals['cash'] - payments_totals['cash'],
        'bank1': receipts_totals['bank1'] - payments_totals['bank1'],
        'bank2': receipts_totals['bank2'] - payments_totals['bank2'],
        'adj': receipts_totals['adj'] - payments_totals['adj'],
    }

    return render(request, 'cash_withdrawals.html', {
        'withdrawals': withdrawals
    })







# from decimal import Decimal
# from django.shortcuts import redirect
# from django.utils import timezone
# from django.contrib import messages

# def cash_transfer(request):
#     if request.method == "POST":
#         from_acc = request.POST.get("from_account")
#         to_acc = request.POST.get("to_account")
#         amount = Decimal(request.POST.get("amount"))

#         if from_acc == to_acc:
#             messages.error(request, "From and To cannot be same")
#             return redirect("cash_withdrawals")

#         # ---- NEGATIVE ENTRY (FROM) ----
#         CashEntry.objects.create(
#             type_of_cash=from_acc,
#             amount=-amount,
#             remarks=f"Transfer to {to_acc}",
#             datetime=timezone.now()
#         )

#         # ---- POSITIVE ENTRY (TO) ----
#         CashEntry.objects.create(
#             type_of_cash=to_acc,
#             amount=amount,
#             remarks=f"Transfer from {from_acc}",
#             datetime=timezone.now()
#         )

#         messages.success(request, "Transfer completed successfully")

#     return redirect("cash_withdrawals")


# from decimal import Decimal
# from django.shortcuts import redirect
# from django.utils import timezone
# from django.contrib import messages
# from django.db.models import Max
# import re

# def cash_transfer(request):
#     if request.method == "POST":

#         from_acc = request.POST.get("from_account")
#         to_acc = request.POST.get("to_account")
#         amount = Decimal(request.POST.get("amount"))

#         if from_acc == to_acc:
#             messages.error(request, "From and To cannot be same")
#             return redirect("cash_withdrawals")

#         # 🔹 Loan prefix (use existing logic / system loan)
#         loan = Loan.objects.first()   # adjust if needed
#         loan_prefix = loan.code if hasattr(loan, "code") else f"LN{loan.id}"

#         # 🔹 Find last TR serial
#         last_tr = CashEntry.objects.filter(
#             code__startswith=f"{loan_prefix}-TR"
#         ).aggregate(max_code=Max("code"))["max_code"]

#         if last_tr:
#             last_num = int(re.findall(r'TR(\d{4})', last_tr)[0])
#             next_num = last_num + 1
#         else:
#             next_num = 1

#         serial = str(next_num).zfill(4)
#         base_code = f"{loan_prefix}-TR{serial}"

#         # ---- FROM (NEGATIVE) ----
#         CashEntry.objects.create(
#             code=f"{base_code}-DR",
#             type_of_cash=from_acc,
#             amount=-amount,
#             remarks=f"Transfer to {to_acc}",
#             datetime=timezone.now()
#         )

#         # ---- TO (POSITIVE) ----
#         CashEntry.objects.create(
#             code=f"{base_code}-CR",
#             type_of_cash=to_acc,
#             amount=amount,
#             remarks=f"Transfer from {from_acc}",
#             datetime=timezone.now()
#         )

#         messages.success(request, f"Transfer successful ({base_code})")

#     return redirect("cash_withdrawals")
from decimal import Decimal
from django.shortcuts import redirect
from django.utils import timezone
from django.contrib import messages
from django.db.models import Max
import re
from .models import CashEntry

def cash_transfer(request):
    if request.method == "POST":

        from_acc = request.POST.get("from_account")
        to_acc = request.POST.get("to_account")
        amount = Decimal(request.POST.get("amount"))

        if from_acc == to_acc:
            messages.error(request, "From and To cannot be same")
            return redirect("cash_withdrawals")

        # 🔹 FIND LAST TR CODE
        last_tr = CashEntry.objects.filter(
            code__startswith="TR"
        ).aggregate(max_code=Max("code"))["max_code"]

        if last_tr:
            last_num = int(re.findall(r'TR(\d{4})', last_tr)[0])
            next_num = last_num + 1
        else:
            next_num = 1

        serial = str(next_num).zfill(4)
        base_code = f"TR{serial}"

        # ---- FROM ACCOUNT (NEGATIVE) ----
        CashEntry.objects.create(
            code=f"{base_code}-DR",
            type_of_cash=from_acc,
            amount=-amount,
            remarks="Transfer",
            datetime=timezone.now(),
            type_of_loan="Transfer"  # <-- set type_of_loan
        )

        # ---- TO ACCOUNT (POSITIVE) ----
        CashEntry.objects.create(
            code=f"{base_code}-CR",
            type_of_cash=to_acc,
            amount=amount,
            remarks="Transfer",
            datetime=timezone.now(),
            type_of_loan="Transfer"  # <-- set type_of_loan
        )

        messages.success(request, f"Transfer completed ({base_code})")

    return redirect("cash_withdrawals")




from django.http import JsonResponse
from decimal import Decimal
from .models import OtherCashTransaction

def others(request):
    if request.method == "POST":
        try:
            gen_no = request.POST.get('gen_no')
            transaction_type = request.POST.get('transaction_type')
            type_of_loan = request.POST.get('type_of_loan')

            if not gen_no or not transaction_type or not type_of_loan:
                return JsonResponse({
                    'status': 'error',
                    'message': 'Missing required fields'
                })

            cash  = Decimal(request.POST.get('cash', '0') or '0')
            bank1 = Decimal(request.POST.get('bank1', '0') or '0')
            bank2 = Decimal(request.POST.get('bank2', '0') or '0')

            amount = cash + bank1 + bank2  # ✅ calculated once

            OtherCashTransaction.objects.create(
                gen_no=gen_no,
                transaction_type=transaction_type,
                type_of_loan=type_of_loan,
                cash=cash,
                bank1=bank1,
                bank2=bank2,
                amount=amount,  # ✅ saved in DB
            )

            return JsonResponse({
                'status': 'success',
                'message': 'Transaction saved successfully'
            })

        except Exception as e:
            print("OTHER CASH ERROR:", e)
            return JsonResponse({
                'status': 'error',
                'message': str(e)
            })

    return JsonResponse({
        'status': 'error',
        'message': 'Invalid request'
    })







# holiday
from decimal import Decimal
from django.shortcuts import render
from django.http import JsonResponse
from .models import OtherCashTransaction, User


from decimal import Decimal
from datetime import datetime
from django.http import JsonResponse
from django.shortcuts import render
from django.utils import timezone
from .models import User, OtherCashTransaction

from decimal import Decimal
from datetime import datetime, time
from django.http import JsonResponse
from django.shortcuts import render
from .models import User, OtherCashTransaction


def others(request):
    gen_no = None
    user_name = None

    # ---------- GET USER ----------
    if request.method == 'POST' and 'gen_no' in request.POST and not request.headers.get('x-requested-with'):
        gen_no = request.POST.get('gen_no')
        user = User.objects.filter(code=gen_no).first()
        user_name = user.name if user else None

    # ---------- ADD RECEIPT / PAYMENT (AJAX) ----------
    if request.method == 'POST' and request.headers.get('x-requested-with') == 'XMLHttpRequest':

        def d(val):
            try:
                return Decimal(val or 0)
            except:
                return Decimal('0')

        data = request.POST

        # ✅ Parse user-selected date (NAIVE datetime for SQLite)
        date_str = data.get('date')
        if not date_str:
            return JsonResponse({'status': 'error', 'message': 'Please select a date'})

        try:
            created_at = datetime.combine(
                datetime.strptime(date_str, "%Y-%m-%d").date(),
                time.min  # 00:00:00
            )
        except ValueError:
            return JsonResponse({'status': 'error', 'message': 'Invalid date format'})

        # ---------- CREATE TRANSACTION ----------
        txn = OtherCashTransaction.objects.create(
            transaction_type=data.get('transaction_type'),  # RECEIPT / PAYMENT
            gen_no=data.get('gen_no'),
            type_of_loan=data.get('type_of_loan'),
            cash=d(data.get('cash')),
            bank1=d(data.get('bank1')),
            bank2=d(data.get('bank2')),
            created_at=created_at  # ✅ SAVES SELECTED DATE
        )

        return JsonResponse({
            'status': 'success',
            'amount': str(txn.amount)
        })

    # ---------- PAGE RENDER ----------
    return render(request, 'others.html', {
        'gen_no': gen_no,
        'user_name': user_name,
    })


from django.http import JsonResponse
from .models import OtherCashTransaction
from decimal import Decimal

def save_other_cash_transaction(request):
    if request.method == "POST":
        gen_no = request.POST.get("gen_no")
        transaction_type = request.POST.get("transaction_type")
        type_of_loan = request.POST.get("type_of_loan")
        cash = request.POST.get("cash") or '0'
        bank1 = request.POST.get("bank1") or '0'
        bank2 = request.POST.get("bank2") or '0'

        try:
            cash = Decimal(cash)
            bank1 = Decimal(bank1)
            bank2 = Decimal(bank2)
        except:
            return JsonResponse({"status":"error", "message":"Invalid input"})

        OtherCashTransaction.objects.create(
            gen_no=gen_no,
            transaction_type=transaction_type,
            type_of_loan=type_of_loan,
            cash=cash,
            bank1=bank1,
            bank2=bank2
            # amount will be calculated automatically in save()
        )
        return JsonResponse({"status":"success"})

    return JsonResponse({"status":"error", "message":"Invalid request"})



# def other_reports_table(request):
#     view_mode = request.GET.get("view", "all")  # all / payments / receipts

#     qs = OtherCashTransaction.objects.all().order_by("-created_at")

#     if view_mode == "payments":
#         qs = qs.filter(transaction_type="PAYMENT")
#     elif view_mode == "receipts":
#         qs = qs.filter(transaction_type="RECEIPT")

#     # map users
#     user_map = {u.code: u.name for u in User.objects.all()}

#     rows = []
#     for t in qs:
#         rows.append({
#             "gen_no": t.gen_no,
#             "name": user_map.get(t.gen_no, "Unknown"),
#             "ref": t.code or "",
#             "cash": t.cash or 0,
#             "bank1": t.bank1 or 0,
#             "bank2": t.bank2 or 0,
#             "adjustment": 0,  # keep 0 unless you add adj later
#             "created_at": t.created_at.strftime("%d-%m-%Y") if t.created_at else "",
#         })

#     return render(request, "reports_list.html", {
#         "rows": rows,
#         "view_mode": view_mode,
#         "report_type": "Other Transactions",
#         "is_loan_or_deposit": False,
#     })


def other_reports_table(request):
    view_mode = request.GET.get("view")  # payments / receipts

    qs = OtherCashTransaction.objects.all().order_by("-created_at")

    # ✅ filter based on card click
    if view_mode == "payments":
        qs = qs.filter(transaction_type="PAYMENT")
        report_type = "Payments"
    elif view_mode == "receipts":
        qs = qs.filter(transaction_type="RECEIPT")
        report_type = "Receipts"
    else:
        report_type = "Other Transactions"

    # ✅ map users (gen_no -> name)
    user_map = {
        u.code: u.name
        for u in User.objects.exclude(code__isnull=True)
    }

    rows = []
    for t in qs:
        rows.append({
            "gen_no": t.gen_no,
            "name": user_map.get(t.gen_no, "Unknown"),
            "ref": t.code or "",
            "cash": t.cash or 0,
            "bank1": t.bank1 or 0,
            "bank2": t.bank2 or 0,
            "adjustment": 0,  # future ready
            "created_at": t.created_at.strftime("%d-%m-%Y") if t.created_at else "",
        })

    return render(request, "other_payments_receipts.html", {
        "rows": rows,
        "report_type": report_type,
    })




# # temp
# from django import forms
# from .models import Loan

# LOAN_TYPE_CHOICES = [
#     ("MTL LOAN", "MTL LOAN"),
#     ("FDL LOAN", "FDL LOAN"),
#     ("KVP/NSC LOAN", "KVP/NSC LOAN"),
#     ("FIXED DEPOSITS", "FIXED DEPOSITS"),
#     ("THRIFT FUNDS", "THRIFT FUNDS"),
#     ("WELFARE COLLECTIONS", "WELFARE COLLECTIONS"),
#     ("SALARY PAID", "SALARY PAID"),
#     ("OFFICE EXPENSES", "OFFICE EXPENSES"),
#     ("OTHER PAYMENTS", "OTHER PAYMENTS"),
#     ("ADMISSION FEES", "ADMISSION FEES"),
#     ("OTHER RECEIPTS", "OTHER RECEIPTS"),
# ]

# class LoanForm(forms.ModelForm):
#     class Meta:
#         model = Loan
#         fields = [
#             "gen_no",
#             "amount",
#             "cash",
#             "bank1",
#             "bank2",
#             "adj",
#             "type_of_loan",
#         ]

#     type_of_loan = forms.ChoiceField(choices=LOAN_TYPE_CHOICES)





# new
# views.py
from django.http import JsonResponse
from .models import OtherCashTransaction

# Fetch existing receipts with code/ref
def fetch_receipts(request):
    gen_no = request.GET.get('gen_no')
    if not gen_no:
        return JsonResponse({'status': 'error', 'message': 'Gen No required'})

    receipts = OtherCashTransaction.objects.filter(
        gen_no=gen_no,
        transaction_type='RECEIPT'
    ).order_by('-created_at', '-id')

    data = [{
        'id': r.id,  # unique ID
        'code': r.code if r.code else f'RC/{r.id}',
        'type_of_loan': r.type_of_loan,
        'cash': str(r.cash),
        'bank1': str(r.bank1),
        'bank2': str(r.bank2),
        'amount': str(r.amount),
        'created_at': r.created_at.strftime('%d-%m-%Y')  # ✅ DD-MM-YYYY
    } for r in receipts]

    return JsonResponse({
        'status': 'success',
        'receipts': data
    })



# Fetch existing payments with code/ref
from django.http import JsonResponse
from .models import OtherCashTransaction

def fetch_payments(request):
    gen_no = request.GET.get('gen_no')
    if not gen_no:
        return JsonResponse({'status': 'error', 'message': 'Gen No required'})

    payments = OtherCashTransaction.objects.filter(
        gen_no=gen_no,
        transaction_type='PAYMENT'
    ).order_by('-created_at', '-id')

    data = [{
        'id': p.id,
        'code': p.code if p.code else f'PM/{p.id}',
        'type_of_loan': p.type_of_loan,
        'cash': str(p.cash),
        'bank1': str(p.bank1),
        'bank2': str(p.bank2),
        'amount': str(p.amount),
        'created_at': r.created_at.strftime('%d-%m-%Y')  # ✅ DD-MM-YYYY
    } for p in payments]

    return JsonResponse({
        'status': 'success',
        'payments': data
    })


# from decimal import Decimal
# from django.shortcuts import render
# from openpyxl import load_workbook
# from .models import User, Loan, OtherCashTransaction, generate_unique_code

# # ---------------- UTILITY FUNCTIONS ----------------
# def clean(v): 
#     return str(v).strip() if v not in [None, "", " "] else ""

# def num(v):
#     try: 
#         return Decimal(v) if v not in [None, "", " "] else Decimal("0.00")
#     except: 
#         return Decimal("0.00")

# # ---------------- UPLOAD EXCEL ----------------
# def upload_excel(request):
#     inserted = {"Users": 0, "Loans": 0, "Deposits": 0, "Others": 0}
#     updated = {"Users": 0, "Loans": 0, "Deposits": 0, "Others": 0}
#     errors = []

#     if request.method == "POST":
#         file = request.FILES.get("excel_file")
#         if not file:
#             return render(request, "upload_excel.html", {"error": "No file selected"})

#         try:
#             wb = load_workbook(file, data_only=True)
#         except Exception as e:
#             return render(request, "upload_excel.html", {"error": f"Invalid file: {e}"})

#         # ================= USERS =================
#         # ================= USERS =================
#         def normalize_gen_no(v):
#             """
#             Accept only real string codes like CN0001
#             Reject Excel numeric junk like 1.0, 2.0
#             """
#             if v is None:
#                 return ""

#             val = str(v).strip()

#             # Reject pure numbers or decimals (Excel junk)
#             if val.replace(".", "").isdigit():
#                 return ""

#             return val

# # ================= USERS =================
#        # ================= USERS =================
#        # ================= USERS =================
# # ================= USERS =================
# # ================= USERS =================
# # ================= USERS =================
#         if "Users" in wb.sheetnames:
#             ws = wb["Users"]
#             headers = [str(c.value).strip().lower().replace(" ", "_") for c in ws[1]]

#             for r, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
#                 data = dict(zip(headers, row))

#                 name = clean(data.get("name"))
#                 raw_gen_no = data.get("gen_no")
#                 mobile = clean(data.get("mobile"))
#                 address = clean(data.get("address"))

#                 if not name:
#                     errors.append(f"Users row {r}: Name is required")
#                     continue

#                 try:
#                     created = False

#                     # -------- NORMALIZE GEN NO ----------
#                     gen_no = ""
#                     if raw_gen_no not in [None, "", " "]:
#                         val = str(raw_gen_no).strip()

#                         # reject Excel junk like 1.0, 2.0
#                         if not val.replace(".", "").isdigit():
#                             gen_no = val

#                     # -------- CASE 1: GEN NO PROVIDED --------
#                     if gen_no:
#                         user = User.objects.filter(code=gen_no).first()

#                         if user:
#                             user.name = name
#                             user.Mobile = mobile
#                             user.Address = address
#                             user.save(update_fields=["name", "Mobile", "Address"])
#                         else:
#                             user = User.objects.create(
#                                 code=gen_no,   # literal save
#                                 name=name,
#                                 Mobile=mobile,
#                                 Address=address
#                             )
#                             created = True

#                     # -------- CASE 2: GEN NO EMPTY --------
#                     else:
#                         user = User.objects.filter(name__iexact=name).order_by("id").first()

#                         if user:
#                             user.Mobile = mobile
#                             user.Address = address
#                             user.save(update_fields=["Mobile", "Address"])
#                         else:
#                             # DO NOT pass code → signal will generate safely
#                             user = User.objects.create(
#                                 name=name,
#                                 Mobile=mobile,
#                                 Address=address
#                             )
#                             created = True

#                     inserted["Users"] += int(created)
#                     updated["Users"] += int(not created)

#                 except Exception as e:
#                     errors.append(f"Users row {r}: {e}")


#         # ================= LOANS =================
#         if "Loans" in wb.sheetnames:
#             ws = wb["Loans"]
#             headers = [str(c.value).strip().lower().replace(" ", "_") for c in ws[1]]

#             for r, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
#                 data = dict(zip(headers, row))
#                 name = clean(data.get("name"))
#                 type_of_loan = clean(data.get("type_of_loan")).upper()
#                 code = clean(data.get("code"))

#                 if not name or not type_of_loan:
#                     continue

#                 if type_of_loan not in ["MTL LOAN", "FDL LOAN", "KVP/NSC LOAN"]:
#                     errors.append(f"Loans row {r}: Invalid loan type {type_of_loan}")
#                     continue

#                 cash = num(data.get("cash"))
#                 bank1 = num(data.get("bank1"))
#                 bank2 = num(data.get("bank2"))
#                 adj = num(data.get("adjustment"))
#                 interest = num(data.get("interest"))
#                 balance = num(data.get("balance"))
#                 amount = cash + bank1 + bank2 + adj

#                 user, _ = User.objects.get_or_create(
#                     name__iexact=name,
#                     defaults={"name": name, "Mobile": "", "Address": ""}
#                 )

#                 payload = {
#                     "gen_no": user.code,
#                     "name": name,
#                     "type_of_loan": type_of_loan,
#                     "amount": amount,
#                     "interest": interest,
#                     "balance": balance,
#                     "cash": str(cash),
#                     "bank1": str(bank1),
#                     "bank2": str(bank2),
#                     "adj": str(adj),
#                     "source": "-",
#                     "loan_status": "Active"
#                 }

#                 try:
#                     if code:
#                         loan, created = Loan.objects.update_or_create(code=code, defaults=payload)
#                     else:
#                         loan = Loan.objects.create(**payload)
#                         created = True

#                     inserted["Loans"] += int(created)
#                     updated["Loans"] += int(not created)
#                 except Exception as e:
#                     errors.append(f"Loans row {r}: {e}")

#         # ================= DEPOSITS =================
#         if "Deposits" in wb.sheetnames or "Loans" in wb.sheetnames:
#             ws = wb["Deposits"] if "Deposits" in wb.sheetnames else wb["Loans"]
#             headers = [str(c.value).strip().lower().replace(" ", "_") for c in ws[1]]

#             for r, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
#                 data = dict(zip(headers, row))
#                 name = clean(data.get("name"))
#                 type_of_loan = clean(data.get("type_of_loan")).upper()
#                 code = clean(data.get("code"))

#                 if not name or not type_of_loan:
#                     continue

#                 if type_of_loan not in ["FIXED DEPOSITS", "THRIFT FUNDS", "WELFARE COLLECTIONS"]:
#                     continue

#                 cash = num(data.get("cash"))
#                 bank1 = num(data.get("bank1"))
#                 bank2 = num(data.get("bank2"))
#                 adj = num(data.get("adjustment"))
#                 interest = num(data.get("interest"))
#                 balance = num(data.get("balance"))
#                 amount = cash + bank1 + bank2 + adj

#                 user, _ = User.objects.get_or_create(
#                     name__iexact=name,
#                     defaults={"name": name, "Mobile": "", "Address": ""}
#                 )

#                 payload = {
#                     "gen_no": user.code,
#                     "name": name,
#                     "type_of_loan": type_of_loan,
#                     "amount": amount,
#                     "interest": interest,
#                     "balance": balance,
#                     "cash": str(cash),
#                     "bank1": str(bank1),
#                     "bank2": str(bank2),
#                     "adj": str(adj),
#                     "source": "-",
#                     "loan_status": "Active"
#                 }

#                 try:
#                     if code:
#                         deposit, created = Loan.objects.update_or_create(code=code, defaults=payload)
#                     else:
#                         deposit = Loan.objects.create(**payload)
#                         created = True

#                     inserted["Deposits"] += int(created)
#                     updated["Deposits"] += int(not created)
#                 except Exception as e:
#                     errors.append(f"Deposits row {r}: {e}")

#         # ================= OTHERS =================
#         if "Others" in wb.sheetnames:
#             ws = wb["Others"]
#             headers = [str(c.value).strip().lower().replace(" ", "_") if c.value else f"unknown_{i}"
#                         for i, c in enumerate(ws[1])]

#             for r, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
#                 if not row or row[0] in [None, "", "Others"]:
#                     continue

#                 data = dict(zip(headers, row))
#                 name = clean(data.get("name"))
#                 transaction_type = clean(data.get("transaction_type")).upper()
#                 type_of_loan = clean(data.get("type_of_loan"))
#                 code = clean(data.get("code"))

#                 cash = num(data.get("cash"))
#                 bank1 = num(data.get("bank1"))
#                 bank2 = num(data.get("bank2"))

#                 amount = cash + bank1 + bank2

#                 if not transaction_type or not type_of_loan:
#                     errors.append(f"Others row {r}: Missing transaction_type or type_of_loan")
#                     continue

#                 user, _ = User.objects.get_or_create(
#                     name__iexact=name,
#                     defaults={"name": name, "Mobile": "", "Address": ""}
#                 )

#                 payload = {
#                     "gen_no": user.code,
#                     "transaction_type": transaction_type,
#                     "type_of_loan": type_of_loan,
#                     "cash": cash,
#                     "bank1": bank1,
#                     "bank2": bank2,
#                 }

#                 try:
#                     if code:
#                         obj, created = OtherCashTransaction.objects.update_or_create(code=code, defaults=payload)
#                     else:
#                         obj = OtherCashTransaction.objects.create(**payload)
#                         created = True

#                     inserted["Others"] += int(created)
#                     updated["Others"] += int(not created)
#                 except Exception as e:
#                     errors.append(f"Others row {r}: {e}")

#         return render(request, "upload_excel.html", {
#             "inserted": inserted,
#             "updated": updated,
#             "errors": errors
#         })

#     return render(request, "upload_excel.html")

from decimal import Decimal
from django.shortcuts import render
from openpyxl import load_workbook
from .models import User, Loan

# ---------------- UTILITY FUNCTIONS ----------------
def clean(v):
    return str(v).strip() if v not in [None, "", " "] else ""

def num(v):
    try:
        return Decimal(v) if v not in [None, "", " "] else Decimal("0.00")
    except:
        return Decimal("0.00")

def normalize_gen_no(v):
    """
    Accepts: 1, 1.0, CN0001
    Returns clean string or ""
    """
    if v in [None, "", " "]:
        return ""

    val = str(v).strip()
    if val.endswith(".0"):
        val = val[:-2]
    return val

# ---------------- UPLOAD EXCEL (USERS + LOANS) ----------------
def upload_excel(request):
    inserted = {"Users": 0, "Loans": 0}
    updated = {"Users": 0, "Loans": 0}
    errors = []

    if request.method == "POST":
        file = request.FILES.get("excel_file")
        if not file:
            return render(request, "upload_excel.html", {"error": "No file selected"})

        wb = load_workbook(file, data_only=True)

        # ================= USERS =================
        if "Users" in wb.sheetnames:
            ws = wb["Users"]
            headers = [str(c.value).strip().lower().replace(" ", "_") for c in ws[1]]

            for r, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                data = dict(zip(headers, row))

                name = clean(data.get("name"))
                gen_no = normalize_gen_no(data.get("gen_no"))
                mobile = clean(data.get("mobile"))
                address = clean(data.get("address"))

                if not name:
                    errors.append(f"Users row {r}: Name is required")
                    continue

                try:
                    # ---------- GEN NO PRESENT ----------
                    if gen_no:
                        user = User.objects.filter(code=gen_no).first()

                        if user:
                            # Update ONLY mutable fields
                            user.name = name
                            user.Mobile = mobile
                            user.Address = address
                            user.save(update_fields=["name", "Mobile", "Address"])
                            updated["Users"] += 1
                        else:
                            # Create new user with fixed code
                            User.objects.create(
                                code=gen_no,
                                name=name,
                                Mobile=mobile,
                                Address=address
                            )
                            inserted["Users"] += 1

                    # ---------- GEN NO MISSING ----------
                    else:
                        # Lookup by name (case-insensitive), get first match
                        user = User.objects.filter(name__iexact=name).order_by("id").first()

                        if user:
                            user.Mobile = mobile
                            user.Address = address
                            user.save(update_fields=["Mobile", "Address"])
                            updated["Users"] += 1
                        else:
                            # Create new → signal generates gen_no
                            User.objects.create(
                                name=name,
                                Mobile=mobile,
                                Address=address
                            )
                            inserted["Users"] += 1

                except Exception as e:
                    errors.append(f"Users row {r}: {e}")

        # ================= LOANS (MTL/FDL/KVP) =================
        if "Loans" in wb.sheetnames:
            ws = wb["Loans"]
            headers = [str(c.value).strip().lower().replace(" ", "_") for c in ws[1]]

            for r, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                # Skip header/junk rows
                if not row or not row[0] or str(row[0]).strip().upper() in ["LOANS", "NAME"]:
                    continue

                data = dict(zip(headers, row))
                name = clean(data.get("name"))
                type_of_loan = clean(data.get("type_of_loan")).upper()
                code = clean(data.get("code"))

                if not name or not type_of_loan:
                    errors.append(f"Loans row {r}: Name or type_of_loan missing")
                    continue

                if type_of_loan not in ["MTL LOAN", "FDL LOAN", "KVP/NSC LOAN"]:
                    errors.append(f"Loans row {r}: Invalid loan type '{type_of_loan}'")
                    continue

                cash = num(data.get("cash"))
                bank1 = num(data.get("bank1"))
                bank2 = num(data.get("bank2"))
                adj = num(data.get("adjustment"))
                interest = num(data.get("interest"))
                balance = num(data.get("balance"))
                amount = cash + bank1 + bank2 + adj

                # ---------------- USER LOOKUP OR CREATE ----------------
                user = User.objects.filter(name__iexact=name).order_by("id").first()
                if not user:
                    # Create new user → signal generates gen_no
                    user = User.objects.create(name=name, Mobile="", Address="")

                # ---------------- PAYLOAD ----------------
                payload = {
                    "gen_no": user.code,
                    "name": name,
                    "type_of_loan": type_of_loan,
                    "amount": amount,
                    "interest": interest,
                    "balance": balance,
                    "cash": str(cash),
                    "bank1": str(bank1),
                    "bank2": str(bank2),
                    "adj": str(adj),
                    "source": "-",
                    "loan_status": "Active",
                }

                try:
                    if code:
                        # Update or create by literal code
                        loan, created = Loan.objects.update_or_create(code=code, defaults=payload)
                    else:
                        # Create new loan → code auto-generated
                        loan = Loan.objects.create(**payload)
                        created = True

                    inserted["Loans"] += int(created)
                    updated["Loans"] += int(not created)

                except Exception as e:
                    errors.append(f"Loans row {r}: {e}")


# ================= DEPOSITS (FIXED DEPOSITS / THRIFT / WELFARE) =================
# ---------------- DEPOSITS ----------------
# ================= DEPOSITS =================
# ================= DEPOSITS =================
# ================= DEBUG DEPOSITS =================
# ================= DEPOSITS =================
# ================= DEPOSITS =================
        DEPOSIT_TYPES_ALLOWED = {
            "FIXED DEPOSITS",
            "THRIFT FUNDS",
            "WELFARE COLLECTIONS",
        }

        if "Deposits" in wb.sheetnames:
            ws = wb["Deposits"]
            headers = [str(c.value).strip().lower().replace(" ", "_") for c in ws[1]]

            for r, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):

                if not row or not row[0]:
                    continue

                data = dict(zip(headers, row))

                name = clean(data.get("name"))
                type_of_loan = clean(data.get("type_of_loan")).upper()
                code = normalize_gen_no(data.get("code"))

                if not name or not type_of_loan:
                    continue

                if type_of_loan not in DEPOSIT_TYPES_ALLOWED:
                    continue

                cash = num(data.get("cash"))
                bank1 = num(data.get("bank1"))
                bank2 = num(data.get("bank2"))
                adj = num(data.get("adjustment"))
                interest = num(data.get("interest"))
                balance = num(data.get("balance"))

                amount = cash + bank1 + bank2 + adj

                # ---------- USER ----------
                user = User.objects.filter(name__iexact=name).order_by("id").first()
                if not user:
                    user = User.objects.create(name=name, Mobile="", Address="")

                payload = {
                    "gen_no": user.code,
                    "name": name,
                    "type_of_loan": type_of_loan,
                    "amount": amount,
                    "interest": interest,
                    "balance": balance,
                    "cash": str(cash),
                    "bank1": str(bank1),
                    "bank2": str(bank2),
                    "adj": str(adj),
                    "source": "RECEIPT",   # ✅ CRITICAL FIX
                    "loan_status": "Active",
                }

                try:
                    if code:
                        Loan.objects.update_or_create(code=code, defaults=payload)
                        updated["Deposits"] += 1
                    else:
                        Loan.objects.create(**payload)
                        inserted["Deposits"] += 1

                except Exception as e:
                    errors.append(f"Deposits row {r}: {e}")



        # ================= OTHERS (RECEIPT / PAYMENT) =================
# ================= OTHERS =================
        if "Others" in wb.sheetnames:
            ws = wb["Others"]

            headers = [
                str(c.value).strip().lower().replace(" ", "_")
                if c.value else ""
                for c in ws[1]
            ]

            for r, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):

                # -------- SKIP EMPTY / JUNK ROWS --------
                if not row or not row[0]:
                    continue

                first_cell = str(row[0]).strip().upper()
                if first_cell in ["OTHERS", "NAME"]:
                    continue

                data = dict(zip(headers, row))

                name = clean(data.get("name"))
                transaction_type = clean(data.get("transaction_type")).upper()
                type_of_loan = clean(data.get("type_of_loan"))
                code = clean(data.get("code"))

                if not name:
                    errors.append(f"Others row {r}: Name is required")
                    continue

                if transaction_type not in ["RECEIPT", "PAYMENT"]:
                    errors.append(
                        f"Others row {r}: Invalid transaction type '{transaction_type}'"
                    )
                    continue

                # -------- USER LOOKUP OR CREATE --------
                user = User.objects.filter(name__iexact=name).order_by("id").first()
                if not user:
                    user = User.objects.create(
                        name=name,
                        Mobile="",
                        Address=""
                    )

                cash = num(data.get("cash"))
                bank1 = num(data.get("bank1"))
                bank2 = num(data.get("bank2"))

                payload = {
                    "gen_no": user.code,
                    "transaction_type": transaction_type,
                    "type_of_loan": type_of_loan,
                    "cash": cash,
                    "bank1": bank1,
                    "bank2": bank2,
                }

                try:
                    # -------- CASE 1: CODE PROVIDED --------
                    if code:
                        obj, created = OtherCashTransaction.objects.update_or_create(
                            code=code,
                            defaults=payload
                        )

                    # -------- CASE 2: CODE EMPTY --------
                    else:
                        # Always create new → RC / PM generated by signal
                        obj = OtherCashTransaction.objects.create(**payload)
                        created = True

                    inserted["Others"] += int(created)
                    updated["Others"] += int(not created)

                except Exception as e:
                    errors.append(f"Others row {r}: {e}")


        return render(request, "upload_excel.html", {
            "inserted": inserted,
            "updated": updated,
            "errors": errors
        })

    return render(request, "upload_excel.html")



# views.py
# views.py
from django.http import JsonResponse
from .models import User

def get_user_info(request):
    gen_no = request.GET.get('gen_no')
    name = request.GET.get('name')

    if gen_no:
        try:
            user = User.objects.get(code=gen_no)
            return JsonResponse({'name': user.name})
        except User.DoesNotExist:
            return JsonResponse({'name': ''})

    if name:
        try:
            user = User.objects.get(name=name)
            return JsonResponse({'gen_no': user.code})
        except User.DoesNotExist:
            return JsonResponse({'gen_no': ''})

    return JsonResponse({})


from django.http import JsonResponse
from .models import User

def get_user_info_autocomplete(request):
    query_gen = request.GET.get('gen_no', '')
    query_name = request.GET.get('name', '')
    users = User.objects.all()

    if query_gen:
        users = users.filter(code__icontains=query_gen)
    if query_name:
        users = users.filter(name__icontains=query_name)

    results = [{'code': u.code, 'name': u.name} for u in users]
    return JsonResponse(results, safe=False)


from django.http import JsonResponse
from .models import User

def get_user_info_autocomplete(request):
    query_gen = request.GET.get('gen_no', '')
    query_name = request.GET.get('name', '')
    users = User.objects.all()
    if query_gen:
        users = users.filter(code__icontains=query_gen)
    if query_name:
        users = users.filter(name__icontains=query_name)
    results = [{'code': u.code, 'name': u.name} for u in users]
    return JsonResponse(results, safe=False)




from django.shortcuts import render
from django.http import JsonResponse
from .models import User, Loan  # replace Loan with your actual Loan model

# Your existing view stays the same (loans_view or whatever)
def loans_vew(request):
    loans = Loan.objects.all()
    users = User.objects.all()
    return render(request, 'add_loan.html', {'loans': loans, 'users': users})

# NEW: API view for Gen No dropdown
def fetch_users_dropdown(request):
    """
    API endpoint to fetch users for dropdown/autocomplete.
    Supports optional query parameter 'q' for filtering by name.
    """
    query = request.GET.get('q', '')
    users = User.objects.filter(name__icontains=query)[:50]  # limit to 50 results
    data = [{"gen_no": user.code, "name": user.name} for user in users]
    return JsonResponse(data, safe=False)

from django.http import JsonResponse
from django.views.decorators.http import require_GET
from .models import User

@require_GET
def user_autocomplete(request):
    term = request.GET.get('term', '')

    qs = User.objects.filter(name__icontains=term) | User.objects.filter(code__icontains=term)

    data = [
        {
            "label": f"{u.code} - {u.name}",
            "value": u.code,
            "name": u.name,
            "code": u.code
        }
        for u in qs[:20]
    ]

    return JsonResponse(data, safe=False)


from django.shortcuts import render, redirect, get_object_or_404
from .models import User

# LIST USERS
from django.db.models import Q
from .models import User

def users(request):
    query = request.GET.get('q')

    users = User.objects.all().order_by('name')

    if query:
        users = users.filter(
            Q(name__icontains=query) |
            Q(Mobile__icontains=query) |
            Q(code__icontains=query)
        )

    return render(request, 'users.html', {'users': users})

# ADD USER
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import User

def adduser(request):
    if request.method == "POST":
        name = request.POST.get('name').strip()
        mobile = request.POST.get('Mobile').strip()
        code = request.POST.get('code')

        # ---- CHECK DUPLICATE USER ----
        if User.objects.filter(Mobile=mobile).exists():
            messages.error(request, "User with this mobile number already exists.")
            return redirect('adduser')

        if User.objects.filter(name__iexact=name, Mobile=mobile).exists():
            messages.error(request, "User already exists.")
            return redirect('adduser')

        if code and User.objects.filter(code=code).exists():
            messages.error(request, "Gen No already exists.")
            return redirect('adduser')

        # ---- CREATE USER ----
        User.objects.create(
            name=name,
            Mobile=mobile,
            code=code,
            Address=request.POST.get('Address')
        )

        messages.success(request, "User added successfully.")
        return redirect('adduser')

    return render(request, 'add_user.html')



# EDIT USER
def edit_user(request, id):
    user = get_object_or_404(User, id=id)

    if request.method == "POST":
        user.code = request.POST.get("code")
        user.name = request.POST.get("name")
        user.Mobile = request.POST.get("Mobile")
        user.Address = request.POST.get("Address")
        user.save()
        return redirect("users")

    return render(request, "edit_user.html", {"user": user})


# DELETE USER
def delete_user(request, id):
    user = get_object_or_404(User, id=id)
    user.delete()
    return redirect("users")




from django.shortcuts import render, redirect
from .models import Loan
from .forms import LoanForm  # We'll create this next

def add_loan_view(request):
    if request.method == 'POST':
        form = LoanForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('loans')  # Redirect to a list page or same page
    else:
        form = LoanForm()
    return render(request, 'add_loan.html', {'form': form})


# from django.shortcuts import render, redirect
# from django.utils import timezone
# from decimal import Decimal
# from datetime import datetime
# from .models import Loan

# def loanadd(request):
#     if request.method == "POST":
#         gen_no = request.POST.get('gen_no', '').strip()
#         name = request.POST.get('name', '').strip()
#         type_of_loan = request.POST.get('type_of_loan', '').strip()
#         cash = request.POST.get('cash', '0')
#         bank1 = request.POST.get('bank1', '0')
#         bank2 = request.POST.get('bank2', '0')
#         adj = request.POST.get('adj', '0')
#         date_input = request.POST.get('date', '')

#         # Safely parse numbers
#         try: cash = Decimal(cash)
#         except: cash = Decimal('0')
#         try: bank1 = Decimal(bank1)
#         except: bank1 = Decimal('0')
#         try: bank2 = Decimal(bank2)
#         except: bank2 = Decimal('0')
#         try: adj = Decimal(adj)
#         except: adj = Decimal('0')

#         amount = cash + bank1 + bank2 + adj

#         # Parse date
#         if date_input:
#             try:
#                 created_at = datetime.strptime(date_input, "%Y-%m-%d")
#             except ValueError:
#                 created_at = timezone.now()
#         else:
#             created_at = timezone.now()

#         # Save only if gen_no and type_of_loan exist
#         if gen_no and type_of_loan:
#             Loan.objects.create(
#                 gen_no=gen_no,
#                 name=name,
#                 type_of_loan=type_of_loan,
#                 amount=amount,
#                 cash=str(cash),
#                 bank1=str(bank1),
#                 bank2=str(bank2),
#                 adj=str(adj),
#                 created_at=created_at,
#                 loan_status='Active'
#             )
#             return redirect('loanadd')  # redirect to same page after submit

#     return render(request, 'loanadd.html')
from django.shortcuts import render, redirect
from django.utils import timezone
from decimal import Decimal, InvalidOperation
from datetime import datetime
from .models import Loan


def loanadd(request):
    if request.method == "POST":
        gen_nos  = request.POST.getlist('gen_no[]')
        names    = request.POST.getlist('name[]')
        types    = request.POST.getlist('type[]')
        cashes   = request.POST.getlist('cash[]')
        bank1s   = request.POST.getlist('bank1[]')
        bank2s   = request.POST.getlist('bank2[]')
        adjs     = request.POST.getlist('adj[]')
        dates    = request.POST.getlist('date[]')

        for i in range(len(gen_nos)):
            gen_no = gen_nos[i].strip()
            type_of_loan = types[i].strip()

            # skip empty rows
            if not gen_no or not type_of_loan:
                continue

            name = names[i].strip() if i < len(names) else ""

            # safe Decimal parsing
            def to_decimal(val):
                try:
                    return Decimal(val)
                except (InvalidOperation, TypeError):
                    return Decimal('0')

            cash  = to_decimal(cashes[i])
            bank1 = to_decimal(bank1s[i])
            bank2 = to_decimal(bank2s[i])
            adj   = to_decimal(adjs[i])

            amount = cash + bank1 + bank2 + adj

            # date parsing
            if dates[i]:
                try:
                    created_at = datetime.strptime(dates[i], "%Y-%m-%d")
                except ValueError:
                    created_at = timezone.now()
            else:
                created_at = timezone.now()

            Loan.objects.create(
                gen_no=gen_no,
                name=name,
                type_of_loan=type_of_loan,
                amount=amount,
                cash=str(cash),
                bank1=str(bank1),
                bank2=str(bank2),
                adj=str(adj),
                created_at=created_at,
                loan_status='Active'
            )

        return redirect('loanadd')

    # GET request – empty page, JS adds first row
    return render(request, 'loanadd.html')



